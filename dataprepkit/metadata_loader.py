"""Metadata-driven orchestrator for dataprepkit dimensions."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import re
from typing import Any, Callable, Dict, Mapping, Sequence, Literal

import logging
import pandas as pd
import uuid
import warnings
from pydantic import BaseModel, ConfigDict, Field, field_validator
from sqlalchemy import inspect, text
from sqlalchemy.engine import Connection, Engine

from dataprepkit.helpers.schema import ensure_schema_exists
from dataprepkit.scd2 import (
    DEFAULT_SYSTEM_COLUMNS,
    EFFECTIVE_DATE_MAX,
    EFFECTIVE_DATE_MIN,
    _compute_row_hash,
    apply_changes,
)
from dataprepkit.storage import archive_dataframe_path


logger = logging.getLogger(__name__)


class ReservedSourceMember(BaseModel):
    source_value: str
    surrogate_key: int
    join_numeric_key: int


DEFAULT_RESERVED_SOURCE_MEMBERS = (
    ReservedSourceMember(
        source_value="NA",
        surrogate_key=-1,
        join_numeric_key=-1,
    ),
)

class DependencyJoin(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    table: str
    schema_name: str | None = Field(default=None, alias="schema")
    how: Literal["left", "inner"] = "left"
    filter_target_current: bool = True
    on: Sequence[Mapping[str, str]]
    select: Mapping[str, str]
    on_missing: Literal["error", "null"] = "error"
    where: Mapping[str, Sequence[str]] = Field(default_factory=dict)


class SchemaHandling(BaseModel):
    mode: Literal["suggest", "evolve"] = "suggest"


class RunPolicy(BaseModel):
    on_table_failure: Literal["continue", "abort"] = "abort"
    on_dependency_failure: Literal["skip_dependents", "abort"] = "skip_dependents"


DEFAULT_DATETIME_INPUT_FORMAT = "%d/%m/%Y %H:%M"


class ColumnSpec(BaseModel):
    type: str | None = None
    nullable: bool = True
    unique: bool = False
    default: str | None = None
    parse_format: str | None = None
    dayfirst: bool = False
    comment: str | None = None


class DimensionMetadata(BaseModel):
    """Defines the metadata required to load a single dimension table."""

    name: str
    target_table: str
    natural_key_cols: Sequence[str]
    natural_key_specs: Mapping[str, ColumnSpec] = Field(default_factory=dict[str, ColumnSpec])
    data_columns: Mapping[str, ColumnSpec]
    surrogate_key: str
    join_numeric_key: str
    filepath: str
    column_renames: Mapping[str, str] = Field(default_factory=dict[str, str])
    reserved_source_members: Sequence[ReservedSourceMember] = Field(
        default_factory=lambda: list(DEFAULT_RESERVED_SOURCE_MEMBERS)
    )
    required_reserved_source_values: Sequence[str] = Field(default_factory=list)
    description: str | None = None
    schema_handling: SchemaHandling = Field(default_factory=SchemaHandling)
    dependencies: Sequence[DependencyJoin] = Field(default_factory=list)
    run_policy: RunPolicy = Field(default_factory=RunPolicy)
    processing_class: Callable[[pd.DataFrame], pd.DataFrame] | None = None
    archive_path: str | None = None
    target_schema: str | None = None
    archive_batch_id: str | None = None

    @field_validator("natural_key_cols")
    def must_define_key_columns(cls, value: Sequence[str]) -> Sequence[str]:
        if not value:
            raise ValueError("Must provide at least one column.")
        return value

    @field_validator("data_columns")
    def must_define_data_columns(cls, value: Mapping[str, ColumnSpec]) -> Mapping[str, ColumnSpec]:
        if not value:
            raise ValueError("Must provide at least one column.")
        return value


ROOT = Path(__file__).resolve().parents[1]
METADATA_REGISTRY: Dict[str, DimensionMetadata] = {}


class DimensionDependencyError(ValueError):
    """Raised when dimension dependency metadata is invalid."""


class CircularDimensionDependencyError(DimensionDependencyError):
    """Raised when registered dimensions contain a dependency cycle."""


def _normalize_bracket_identifier(name: str | None) -> str | None:
    if name is None:
        return None
    stripped = name.strip()
    if stripped.startswith("[") and stripped.endswith("]"):
        return stripped[1:-1].replace("]]", "]")
    return stripped


def _quote_identifier(engine: Engine, identifier: str) -> str:
    normalized = _normalize_bracket_identifier(identifier) or ""
    if engine.dialect.name == "mssql":
        return f"[{normalized.replace(']', ']]')}]"
    return f'"{normalized.replace(chr(34), chr(34) * 2)}"'


def _safe_identifier_token(identifier: str) -> str:
    normalized = (_normalize_bracket_identifier(identifier) or "").strip('"')
    token = re.sub(r"[^0-9A-Za-z_]+", "_", normalized).strip("_")
    return token or "col"


def _normalize_column_specs(
    raw: Sequence[str] | Mapping[str, Any]
) -> Mapping[str, ColumnSpec]:
    if isinstance(raw, Mapping):
        normalized = {}
        for name, spec in raw.items():
            if isinstance(spec, ColumnSpec):
                normalized[name] = spec
            else:
                normalized[name] = ColumnSpec(**spec)
        return normalized
    return {
        name: ColumnSpec(type=None, nullable=False)
        for name in raw
    }


def register_metadata(
    name: str,
    metadata: Dict[str, object],
    *,
    metadata_registry: Dict[str, DimensionMetadata] | None = None,
    archive_base_dir: str | None = None,
    archive_batch_id: str | None = None,
    batch_id: str | None = None,
) -> None:
    """Register metadata using a JSON-like dictionary for familiarity with old_code.py."""
    metadata = metadata.copy()
    schema = metadata.pop("target_schema", None)
    archive_config = metadata.pop("archive_config", None)
    resolved_batch_id = archive_batch_id or batch_id
    if not archive_config and archive_base_dir and resolved_batch_id:
        archive_config = {"base_dir": archive_base_dir, "batch_id": resolved_batch_id}
    if resolved_batch_id:
        metadata["archive_batch_id"] = resolved_batch_id
    if schema is None:
        schema = metadata.pop("schema", None)
    if "data_columns" in metadata:
        metadata["data_columns"] = _normalize_column_specs(
            metadata["data_columns"]  # type: ignore[arg-type]
        )
    if "natural_key_specs" in metadata:
        metadata["natural_key_specs"] = _normalize_column_specs(
            metadata["natural_key_specs"]  # type: ignore[arg-type]
        )
    if schema:
        table = metadata.get("target_table") or ""
        if schema and table and "." not in table:
            metadata["target_table"] = f"{schema}.{table}"
        metadata["target_schema"] = schema
    if archive_config:
        target_table = metadata.get("target_table", "")
        path_info = archive_dataframe_path(
            table_name=target_table,
            batch_id=archive_config.get("batch_id"),
            base_dir=archive_config.get("base_dir", ""),
        )
        metadata["archive_path"] = path_info.file_path
    registry = metadata_registry if metadata_registry is not None else METADATA_REGISTRY
    registry[name] = DimensionMetadata(name=name, **metadata)


def _register_default_metadata() -> None:
    sample_path = ROOT / "examples" / "dummy_dimension.csv"
    register_metadata(
        "dummy_dimension",
        {
            "target_table": "dimension",
            "natural_key_cols": ["natural_key"],
            "data_columns": {"data_column": {"type": "TEXT"}},
            "surrogate_key": "surrogate_key",
            "join_numeric_key": "join_numeric_key",
            "filepath": str(sample_path),
            "description": "Placeholder dimension used in demos.",
        },
    )


def _read_excel_one_sheet_openpyxl(filepath: str) -> pd.DataFrame:
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=True, read_only=True)

    if len(wb.sheetnames) != 1:
        raise ValueError(f"Expected exactly one sheet, found {len(wb.sheetnames)}")

    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]

    df = pd.DataFrame(data, columns=header)
    df = df.dropna(how="all")
    return _normalize_input_column_names(df)


def _normalize_input_column_names(df: pd.DataFrame) -> pd.DataFrame:
    normalized = [str(column).strip() for column in df.columns]
    if len(normalized) != len(set(normalized)):
        raise ValueError("Input data contains duplicate column names after trimming whitespace.")
    renamed = df.copy()
    renamed.columns = normalized
    return renamed


def _default_csv_reader(filepath: str) -> pd.DataFrame:
    if filepath.lower().endswith(".xlsx"):
        return _read_excel_one_sheet_openpyxl(filepath)
    if filepath.lower().endswith(".parquet"):
        return _normalize_input_column_names(pd.read_parquet(filepath))

    return _normalize_input_column_names(
        pd.read_csv(
            filepath,
            header=0,
            encoding="utf-8",
            dtype=str,
            keep_default_na=False,
            na_values=["NULL"],
        )
    )

def get_metadata(
    name: str,
    metadata_registry: Mapping[str, DimensionMetadata] | None = None,
) -> DimensionMetadata:
    """Return metadata record by name."""
    registry = metadata_registry if metadata_registry is not None else METADATA_REGISTRY
    try:
        return registry[name]
    except KeyError as exc:
        raise KeyError(f"Unknown metadata entry: '{name}'") from exc


def build_dimension_dependency_graph(
    metadata_registry: Mapping[str, DimensionMetadata] | None = None,
) -> dict[str, set[str]]:
    registry = metadata_registry or METADATA_REGISTRY
    table_index = _build_dimension_table_index(registry)
    graph: dict[str, set[str]] = {name: set() for name in registry}
    for name, metadata in registry.items():
        for dependency in metadata.dependencies:
            dependency_name = _resolve_dimension_dependency_name(
                dependency,
                table_index,
            )
            if dependency_name is not None:
                graph[name].add(dependency_name)
    return graph


def resolve_dimension_execution_order(
    metadata_registry: Mapping[str, DimensionMetadata] | None = None,
) -> list[str]:
    graph = build_dimension_dependency_graph(metadata_registry)
    pending = {name: set(dependencies) for name, dependencies in graph.items()}
    order: list[str] = []
    ready = sorted(name for name, dependencies in pending.items() if not dependencies)

    while ready:
        current = ready.pop(0)
        order.append(current)
        pending.pop(current, None)
        newly_ready = []
        for name, dependencies in pending.items():
            if current in dependencies:
                dependencies.remove(current)
                if not dependencies:
                    newly_ready.append(name)
        if newly_ready:
            ready.extend(sorted(newly_ready))
            ready.sort()

    if pending:
        cycle = _find_dimension_dependency_cycle(graph)
        cycle_text = " -> ".join(cycle) if cycle else ", ".join(sorted(pending))
        raise CircularDimensionDependencyError(
            f"Circular dimension dependencies detected: {cycle_text}"
        )

    return order


def run_dimensions_in_dependency_order(
    engine: Engine,
    names: Sequence[str] | None = None,
    metadata_registry: Mapping[str, DimensionMetadata] | None = None,
    **run_dimension_kwargs,
) -> list[pd.DataFrame]:
    registry = metadata_registry or METADATA_REGISTRY
    selected_registry = (
        {name: registry[name] for name in names}
        if names is not None
        else registry
    )
    execution_order = resolve_dimension_execution_order(selected_registry)
    return [
        run_dimension(
            engine,
            name,
            metadata_registry=selected_registry,
            **run_dimension_kwargs,
        )
        for name in execution_order
    ]


def _split_reserved_member_rows(
    metadata: DimensionMetadata,
    incoming: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, pd.Series]]:
    if not metadata.natural_key_cols:
        return incoming, {}
    remaining = incoming.copy()
    reserved_rows: dict[str, pd.Series] = {}
    required_values = set(metadata.required_reserved_source_values)
    for member in metadata.reserved_source_members:
        mask = _reserved_member_mask(
            remaining,
            metadata.natural_key_cols,
            member.source_value,
        )
        matches = int(mask.sum())
        if matches == 0:
            if member.source_value in required_values:
                natural_key_values = {
                    column: member.source_value
                    for column in metadata.natural_key_cols
                }
                raise ValueError(
                    "Reserved member missing from input for "
                    f"{metadata.name}: surrogate_key={member.surrogate_key}, "
                    f"natural_key_values={natural_key_values}"
                )
            continue
        if matches > 1:
            natural_key_values = {
                column: member.source_value
                for column in metadata.natural_key_cols
            }
            raise ValueError(
                "Reserved member appears multiple times in input for "
                f"{metadata.name}: surrogate_key={member.surrogate_key}, "
                f"natural_key_values={natural_key_values}"
            )
        reserved_rows[member.source_value] = remaining.loc[mask].iloc[0].copy()
        remaining = remaining.loc[~mask].copy()
    return remaining, reserved_rows


def _reserved_member_mask(
    incoming: pd.DataFrame,
    natural_key_cols: Sequence[str],
    source_value: str,
) -> pd.Series:
    mask = pd.Series(True, index=incoming.index)
    for column in natural_key_cols:
        mask &= incoming[column] == source_value
    return mask


def _upsert_reserved_member(
    engine: Engine,
    metadata: DimensionMetadata,
    member: ReservedSourceMember,
    row: pd.Series,
    data_columns: Sequence[str],
    execution_time: str,
    has_batch_id: bool,
    has_archive_filename: bool,
    batch_id: str,
    archive_filename: str,
) -> None:
    row_hash = _compute_row_hash(row, data_columns)
    natural_key_values = {
        column: _normalize_value_for_sql(row[column])
        for column in metadata.natural_key_cols
    }
    data_values = {
        column: _normalize_value_for_sql(row.get(column))
        for column in data_columns
    }

    with engine.begin() as conn:
        quote = lambda value: _quote_identifier(engine, value)
        existing = conn.execute(
            text(
                f"""
                SELECT {quote(metadata.surrogate_key)}, {quote(metadata.join_numeric_key)}, {quote(DEFAULT_SYSTEM_COLUMNS['insert_date'])}
                FROM {metadata.target_table}
                WHERE {quote(metadata.surrogate_key)} = :surrogate_key
                """
            ),
            {"surrogate_key": member.surrogate_key},
        ).fetchone()
        _raise_on_reserved_member_conflict(conn, metadata, member)
        join_numeric_key = member.join_numeric_key
        if existing is None:
            _insert_reserved_member(
                conn=conn,
                metadata=metadata,
                member=member,
                natural_key_values=natural_key_values,
                data_values=data_values,
                row_hash=row_hash,
                join_numeric_key=join_numeric_key,
                execution_time=execution_time,
                has_batch_id=has_batch_id,
                has_archive_filename=has_archive_filename,
                batch_id=batch_id,
                archive_filename=archive_filename,
            )
            return
        _update_reserved_member(
            conn=conn,
            metadata=metadata,
            member=member,
            natural_key_values=natural_key_values,
            data_values=data_values,
            row_hash=row_hash,
            join_numeric_key=join_numeric_key,
            execution_time=execution_time,
            insert_date=existing[2] or execution_time,
            has_batch_id=has_batch_id,
            has_archive_filename=has_archive_filename,
            batch_id=batch_id,
            archive_filename=archive_filename,
        )


def _raise_on_reserved_member_conflict(
    conn: Connection,
    metadata: DimensionMetadata,
    member: ReservedSourceMember,
) -> None:
    quote = lambda value: _quote_identifier(conn.engine, value)
    predicates = [
        f"{quote(column)} = :nk_{index}"
        for index, column in enumerate(metadata.natural_key_cols)
    ]
    params = {
        f"nk_{index}": member.source_value
        for index, column in enumerate(metadata.natural_key_cols)
    }
    params["surrogate_key"] = member.surrogate_key
    conflict = conn.execute(
        text(
            f"""
            SELECT {quote(metadata.surrogate_key)}
            FROM {metadata.target_table}
            WHERE {' AND '.join(predicates)}
              AND {quote(metadata.surrogate_key)} != :surrogate_key
              AND {quote(DEFAULT_SYSTEM_COLUMNS['current_ind'])} = 1
            """
        ),
        params,
    ).fetchone()
    if conflict is not None:
        raise RuntimeError(
            "Reserved member conflicts with an existing current row in "
            f"{metadata.target_table}: natural_key_values="
            f"{ {column: member.source_value for column in metadata.natural_key_cols} }, "
            f"existing_surrogate_key={conflict[0]}, reserved_surrogate_key={member.surrogate_key}"
        )


def _insert_reserved_member(
    conn: Connection,
    metadata: DimensionMetadata,
    member: ReservedSourceMember,
    natural_key_values: Mapping[str, Any],
    data_values: Mapping[str, Any],
    row_hash: str,
    join_numeric_key: int,
    execution_time: str,
    has_batch_id: bool,
    has_archive_filename: bool,
    batch_id: str,
    archive_filename: str,
) -> None:
    column_values = [
        (metadata.surrogate_key, member.surrogate_key),
        *[(column, natural_key_values[column]) for column in metadata.natural_key_cols],
        (metadata.join_numeric_key, join_numeric_key),
        *[(column, data_values[column]) for column in data_values],
        (DEFAULT_SYSTEM_COLUMNS["row_hash"], row_hash),
        (DEFAULT_SYSTEM_COLUMNS["insert_date"], execution_time),
        (DEFAULT_SYSTEM_COLUMNS["update_date"], None),
        (DEFAULT_SYSTEM_COLUMNS["effective_date_start"], EFFECTIVE_DATE_MIN),
        (DEFAULT_SYSTEM_COLUMNS["effective_date_end"], EFFECTIVE_DATE_MAX),
        (DEFAULT_SYSTEM_COLUMNS["current_ind"], 1),
        (DEFAULT_SYSTEM_COLUMNS["deleted_ind"], 0),
    ]
    if has_batch_id:
        column_values.append((DEFAULT_SYSTEM_COLUMNS["batch_id"], batch_id))
    if has_archive_filename:
        column_values.append((DEFAULT_SYSTEM_COLUMNS["archive_filename"], archive_filename))
    columns: list[str] = []
    placeholders: list[str] = []
    params: dict[str, Any] = {}
    for index, (column, value) in enumerate(column_values):
        param_name = f"insert_value_{index}"
        columns.append(_quote_identifier(conn.engine, column))
        placeholders.append(f":{param_name}")
        params[param_name] = value
    insert_sql = text(
        f"""
        INSERT INTO {metadata.target_table} ({', '.join(columns)})
        VALUES ({', '.join(placeholders)})
        """
    )
    _execute_identity_insert(
        conn,
        metadata.target_table,
        insert_sql,
        params,
        enabled=conn.engine.dialect.name == "mssql",
    )


def _update_reserved_member(
    conn: Connection,
    metadata: DimensionMetadata,
    member: ReservedSourceMember,
    natural_key_values: Mapping[str, Any],
    data_values: Mapping[str, Any],
    row_hash: str,
    join_numeric_key: int,
    execution_time: str,
    insert_date: str,
    has_batch_id: bool,
    has_archive_filename: bool,
    batch_id: str,
    archive_filename: str,
) -> None:
    value_items = [
        (metadata.join_numeric_key, join_numeric_key),
        (DEFAULT_SYSTEM_COLUMNS["row_hash"], row_hash),
        (DEFAULT_SYSTEM_COLUMNS["insert_date"], insert_date),
        (DEFAULT_SYSTEM_COLUMNS["update_date"], None),
        (DEFAULT_SYSTEM_COLUMNS["effective_date_start"], EFFECTIVE_DATE_MIN),
        (DEFAULT_SYSTEM_COLUMNS["effective_date_end"], EFFECTIVE_DATE_MAX),
        (DEFAULT_SYSTEM_COLUMNS["current_ind"], 1),
        (DEFAULT_SYSTEM_COLUMNS["deleted_ind"], 0),
        *[(column, natural_key_values[column]) for column in natural_key_values],
        *[(column, data_values[column]) for column in data_values],
    ]
    if has_batch_id:
        value_items.append((DEFAULT_SYSTEM_COLUMNS["batch_id"], batch_id))
    if has_archive_filename:
        value_items.append((DEFAULT_SYSTEM_COLUMNS["archive_filename"], archive_filename))
    set_clauses: list[str] = []
    params: dict[str, Any] = {"reserved_surrogate_key": member.surrogate_key}
    for index, (column, value) in enumerate(value_items):
        param_name = f"update_value_{index}"
        set_clauses.append(f"{_quote_identifier(conn.engine, column)} = :{param_name}")
        params[param_name] = value
    conn.execute(
        text(
            f"""
            UPDATE {metadata.target_table}
            SET {', '.join(set_clauses)}
            WHERE {_quote_identifier(conn.engine, metadata.surrogate_key)} = :reserved_surrogate_key
            """
        ),
        params,
    )


def _execute_identity_insert(
    conn: Connection,
    table_name: str,
    statement,
    params: Mapping[str, Any],
    enabled: bool,
) -> None:
    if not enabled:
        conn.execute(statement, params)
        return
    conn.execute(text(f"SET IDENTITY_INSERT {table_name} ON"))
    try:
        conn.execute(statement, params)
    finally:
        conn.execute(text(f"SET IDENTITY_INSERT {table_name} OFF"))


def _normalize_value_for_sql(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except (AttributeError, ValueError, TypeError):
            pass
    return value


def _get_reserved_member(
    metadata: DimensionMetadata,
    source_value: str,
) -> ReservedSourceMember:
    for member in metadata.reserved_source_members:
        if member.source_value == source_value:
            return member
    raise ValueError(
        f"Reserved source value '{source_value}' is not configured for {metadata.name}"
    )


def _delete_reserved_member(
    engine: Engine,
    metadata: DimensionMetadata,
    source_value: str,
) -> int:
    inspector = inspect(engine)
    schema, table = _split_table_name(metadata.target_table)
    if not inspector.has_table(table, schema=schema):
        return 0
    params = {
        f"nk_{index}": source_value
        for index, _column in enumerate(metadata.natural_key_cols)
    }
    predicate = " AND ".join(
        f"{_quote_identifier(engine, column)} = :nk_{index}"
        for index, column in enumerate(metadata.natural_key_cols)
    )
    with engine.begin() as conn:
        result = conn.execute(
            text(f"DELETE FROM {metadata.target_table} WHERE {predicate}"),
            params,
        )
    return result.rowcount or 0




def run_dimension(
    engine: Engine,
    metadata_name: str,
    *,
    metadata_registry: Mapping[str, DimensionMetadata] | None = None,
    override_df: pd.DataFrame | None = None,
    csv_reader: callable = _default_csv_reader,
    staging_use_openrowset_parquet: bool = False,
    staging_parquet_base_dir: str | None = None,
    staging_copy_source_base_url: str | None = None,
    staging_copy_into_options: str = "",
) -> pd.DataFrame:
    """
    Load a dimension based on metadata and apply SCD2 semantics.

    Parameters
    ----------
    engine
        SQLAlchemy engine.
    metadata_name
        Identifier in the metadata registry.
    override_df
        Optional DataFrame to bypass file reading (useful for tests).
    csv_reader
        Callable to read the source file (defaults to pandas.read_csv).
    staging_use_openrowset_parquet
        If True (MSSQL only), stage incoming snapshot via parquet + OPENROWSET load.
    staging_parquet_base_dir
        Base directory where staging parquet files are written.
    staging_copy_source_base_url
        Optional SQL-visible base path/URI for OPENROWSET BULK source.
    staging_copy_into_options
        Optional OPENROWSET options suffix (for example ", MAXERRORS = 10").
    """
    metadata = get_metadata(metadata_name, metadata_registry=metadata_registry)
    incoming = (
        override_df.copy()
        if override_df is not None
        else csv_reader(metadata.filepath)
    )
    archive_snapshot = incoming.copy()
    logger.info(
        "Read raw snapshot for %s from %s (%d rows, columns=%s)",
        metadata_name,
        metadata.filepath,
        len(incoming),
        list(incoming.columns),
    )
    incoming = _cast_data_columns(incoming, metadata)
    _ensure_target_table(engine, metadata)
    _apply_table_description(engine, metadata)
    if metadata.column_renames:
        incoming = incoming.rename(columns=metadata.column_renames)
        if incoming.columns.duplicated().any():
            logger.error(
                "Column rename collision detected with %s",
                metadata.column_renames,
            )
            raise ValueError(
                "Column renames introduced duplicate column names; check metadata."
            )
    if metadata.processing_class:
        incoming = metadata.processing_class(incoming)
        logger.debug(
            "Applied processing_class for %s; first rows:\n%s",
            metadata_name,
            incoming.head(),
        )
    incoming = _apply_dependency_joins(incoming, metadata.dependencies, engine)
    incoming = _cast_data_columns(incoming, metadata)
    logger.debug(
        "After dependency joins for %s: %d rows, columns=%s",
        metadata_name,
        len(incoming),
        list(incoming.columns),
    )
    execution_time = _capture_execution_time()
    execution_time_iso = execution_time.isoformat(timespec="milliseconds")
    archive_dest = _resolve_archive_destination(metadata, execution_time)
    archive_filename = archive_dest.name if archive_dest else Path(metadata.filepath).name
    metadata_batch_id = metadata.archive_batch_id or metadata.name
    logger.info(
        "Loaded dimension '%s' from %s (%d rows)",
        metadata.name,
        metadata.filepath,
        len(incoming),
    )
    logger.info("Execution timestamp: %s", execution_time_iso)
    available_columns = _get_target_columns(engine, metadata.target_table)
    safe_data_columns, missing = _handle_schema_drift(
        engine,
        metadata,
        available_columns,
        metadata.data_columns,
    )
    has_batch_id = DEFAULT_SYSTEM_COLUMNS["batch_id"] in available_columns
    has_archive_filename = DEFAULT_SYSTEM_COLUMNS["archive_filename"] in available_columns
    system_columns = DEFAULT_SYSTEM_COLUMNS
    if missing:
        logger.warning(
            "Schema drift for target %s: missing columns %s; using safe write set %s",
            metadata.target_table,
            missing,
            safe_data_columns,
        )
    incoming, reserved_rows = _split_reserved_member_rows(
        metadata=metadata,
        incoming=incoming,
    )

    rows_processed = len(incoming)
    logger.info(
        "Starting table %s: rows=%d, execution_time=%s",
        metadata.target_table,
        rows_processed,
        execution_time,
    )
    start_ts = datetime.now(timezone.utc)
    changes_applied = False
    try:
            changes_applied = apply_changes(
                engine=engine,
                target_table=metadata.target_table,
                incoming=incoming,
                natural_key_cols=list(metadata.natural_key_cols),
                data_cols=safe_data_columns,
                join_numeric_key_col=metadata.join_numeric_key,
                surrogate_key_col=metadata.surrogate_key,
                system_columns=system_columns,
                nullable_columns=[
                    name
                    for name, spec in metadata.data_columns.items()
                    if spec.nullable
                ],
                execution_time=execution_time_iso,
                batch_id=metadata_batch_id,
                archive_filename=archive_filename,
                has_batch_id=has_batch_id,
                has_archive_filename=has_archive_filename,
                staging_use_openrowset_parquet=staging_use_openrowset_parquet,
                staging_parquet_base_dir=staging_parquet_base_dir,
                staging_copy_source_base_url=staging_copy_source_base_url,
                staging_copy_into_options=staging_copy_into_options,
            )
    except Exception as exc:
        duration = (datetime.now(timezone.utc) - start_ts).total_seconds()
        logger.error("SCD2 invocation failed for %s: %s", metadata.name, exc)
        logger.info("Run policy on table failure: %s", metadata.run_policy.on_table_failure)
        logger.info(
            "Table %s failed after %.3fs (rows=%d)",
            metadata.target_table,
            duration,
            rows_processed,
        )
        if metadata.run_policy.on_table_failure == "abort":
            raise
        logger.info("Continuing despite table failure per policy.")
        return incoming
    else:
        required_reserved_values = set(metadata.required_reserved_source_values)
        for member in metadata.reserved_source_members:
            reserved_row = reserved_rows.get(member.source_value)
            if reserved_row is not None:
                _upsert_reserved_member(
                    engine=engine,
                    metadata=metadata,
                    member=member,
                    row=reserved_row,
                    data_columns=safe_data_columns,
                    execution_time=execution_time_iso,
                    has_batch_id=has_batch_id,
                    has_archive_filename=has_archive_filename,
                    batch_id=metadata_batch_id,
                    archive_filename=archive_filename,
                )
                changes_applied = True
            elif member.source_value not in required_reserved_values:
                changes_applied = bool(
                    _delete_reserved_member(
                        engine,
                        metadata,
                        member.source_value,
                    )
                ) or changes_applied
        duration = (datetime.now(timezone.utc) - start_ts).total_seconds()
        logger.info("Run policy on success: %s", metadata.run_policy.on_table_failure)
        logger.info(
            "Table %s completed in %.3fs (rows=%d)",
            metadata.target_table,
            duration,
            rows_processed,
        )
        if archive_dest:
            if {"batch_id", "archive_filename"} <= system_columns.keys():
                if changes_applied:
                    _archive_snapshot(archive_snapshot, metadata, archive_dest)
            else:
                missing_cols = {"batch_id", "archive_filename"} - set(system_columns.keys())
                logger.warning(
                    "Failed to archive snapshot to %s: missing system columns %s",
                    archive_dest,
                    missing_cols,
                )
    logger.info("SCD2 classification counts: not available")
    _post_scd2_validation(engine, metadata.target_table, metadata.natural_key_cols)
    return incoming


def run_dimension_copy_into(
    engine: Engine,
    metadata_name: str,
    *,
    metadata_registry: Mapping[str, DimensionMetadata] | None = None,
    destination_table: str,
    copy_source_base_url: str,
    parquet_base_dir: str,
    copy_into_options: str = "",
    override_df: pd.DataFrame | None = None,
    csv_reader: callable = _default_csv_reader,
) -> str:
    """
    Experimental path: read via pandas, persist parquet, then invoke COPY INTO.

    Returns the external source URL used in the COPY INTO command.
    """
    metadata = get_metadata(metadata_name, metadata_registry=metadata_registry)
    incoming = (
        override_df.copy()
        if override_df is not None
        else csv_reader(metadata.filepath)
    )
    incoming = _cast_data_columns(incoming, metadata)
    if metadata.column_renames:
        incoming = incoming.rename(columns=metadata.column_renames)
        if incoming.columns.duplicated().any():
            raise ValueError(
                "Column renames introduced duplicate column names; check metadata."
            )
    if metadata.processing_class:
        incoming = metadata.processing_class(incoming)
    incoming = _apply_dependency_joins(incoming, metadata.dependencies, engine)
    incoming = _cast_data_columns(incoming, metadata)

    _, target_table = _split_table_name(metadata.target_table)
    batch_id = metadata.archive_batch_id or metadata.name
    parquet_info = archive_dataframe_path(
        table_name=target_table,
        batch_id=batch_id,
        base_dir=parquet_base_dir,
    )
    parquet_path = Path(parquet_info.file_path)
    incoming.to_parquet(parquet_path, index=False)

    source_url = (
        f"{copy_source_base_url.rstrip('/')}/{target_table}/{parquet_path.name}"
    )
    options_sql = copy_into_options.strip()
    if options_sql and not options_sql.startswith(","):
        options_sql = f", {options_sql}"

    copy_sql = text(
        f"""
        COPY INTO {destination_table}
        FROM :source_url
        WITH (
            FILE_TYPE = 'PARQUET'{options_sql}
        )
        """
    )
    with engine.begin() as conn:
        conn.execute(copy_sql, {"source_url": source_url})
    return source_url


def _get_target_columns(engine: Engine, table_name: str) -> set[str]:
    inspector = inspect(engine)
    schema, table = _split_table_name(table_name)
    if not inspector.has_table(table, schema=schema):
        raise RuntimeError(f"Target table '{table_name}' does not exist.")
    return {col["name"] for col in inspector.get_columns(table, schema=schema)}


def _column_spec_clause(name: str, spec: ColumnSpec, engine: Engine) -> str:
    column_type = spec.type or _column_type_for_engine(engine)
    constraints = []
    if not spec.nullable:
        constraints.append("NOT NULL")
    if spec.unique:
        constraints.append("UNIQUE")
    if spec.default is not None:
        constraints.append(f"DEFAULT {spec.default}")
    constraint_sql = " ".join(constraints)
    return f"{_quote_identifier(engine, name)} {column_type} {constraint_sql}".strip()


def _ensure_target_table(engine: Engine, metadata: DimensionMetadata) -> None:
    inspector = inspect(engine)
    schema_name, table_name = _split_table_name(metadata.target_table)
    ensure_schema_exists(engine, schema_name)
    if inspector.has_table(table_name, schema=schema_name):
        current_columns = {col["name"] for col in inspector.get_columns(table_name, schema=schema_name)}
        expected_columns = _expected_column_names(metadata)
        missing = expected_columns - current_columns
        if missing:
            if metadata.schema_handling.mode == "evolve":
                logger.info(
                    "Existing table '%s' is missing metadata columns %s; schema handling=%s",
                    metadata.target_table,
                    missing,
                    metadata.schema_handling.mode,
                )
                _evolve_table_columns(engine, metadata, missing)
            else:
                logger.warning(
                    "Existing table '%s' is missing metadata columns %s; schema handling=%s",
                    metadata.target_table,
                    missing,
                    metadata.schema_handling.mode,
                )
        return
    natural_specs = {
        **{col: spec for col, spec in metadata.natural_key_specs.items()},
        **{
            col: ColumnSpec(type=None, nullable=False)
            for col in metadata.natural_key_cols
            if col not in metadata.natural_key_specs
        },
    }
    column_defs = [
        _column_spec_clause(name, spec, engine)
        for name, spec in natural_specs.items()
    ]
    column_defs += [
        _column_spec_clause(name, spec, engine)
        for name, spec in metadata.data_columns.items()
    ]
    system_columns = [
        f"{_quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS['row_hash'])} {_system_column_type(DEFAULT_SYSTEM_COLUMNS['row_hash'], engine)} NOT NULL",
        f"{_quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS['insert_date'])} {_system_column_type(DEFAULT_SYSTEM_COLUMNS['insert_date'], engine)} NOT NULL",
        f"{_quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS['update_date'])} {_system_column_type(DEFAULT_SYSTEM_COLUMNS['update_date'], engine)}",
        f"{_quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS['effective_date_start'])} {_system_column_type(DEFAULT_SYSTEM_COLUMNS['effective_date_start'], engine)} NOT NULL",
        f"{_quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS['effective_date_end'])} {_system_column_type(DEFAULT_SYSTEM_COLUMNS['effective_date_end'], engine)} NOT NULL",
        f"{_quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS['current_ind'])} {_system_column_type(DEFAULT_SYSTEM_COLUMNS['current_ind'], engine)} NOT NULL",
        f"{_quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS['deleted_ind'])} {_system_column_type(DEFAULT_SYSTEM_COLUMNS['deleted_ind'], engine)} NOT NULL",
        f"{_quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS['batch_id'])} {_system_column_type(DEFAULT_SYSTEM_COLUMNS['batch_id'], engine)} NOT NULL",
        f"{_quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS['archive_filename'])} {_system_column_type(DEFAULT_SYSTEM_COLUMNS['archive_filename'], engine)} NOT NULL",
    ]
    surrogate_clause = _surrogate_column_clause(engine, metadata.surrogate_key)
    join_numeric_clause = _join_numeric_clause(engine, metadata.join_numeric_key)
    create_sql = f"""
        CREATE TABLE {metadata.target_table} (
            {surrogate_clause},
            {join_numeric_clause},
            {', '.join(column_defs + system_columns)}
        )
        """
    logger.info("Creating target table %s with DDL:\n%s", metadata.target_table, create_sql)
    with engine.begin() as conn:
        conn.execute(text(create_sql))
        conn.execute(
            text(
                _current_business_key_index_sql(
                    metadata.target_table,
                    metadata.natural_key_cols,
                    DEFAULT_SYSTEM_COLUMNS["current_ind"],
                )
            )
        )


def _split_table_name(name: str) -> tuple[str | None, str]:
    if "." in name:
        schema, table = name.split(".", 1)
    else:
        return None, name
    schema = schema.strip("[]\"")
    table = table.strip("[]\"")
    return schema, table


def _current_business_key_index_sql(
    target_table: str, natural_key_cols: Sequence[str], current_ind: str
) -> str:
    _, table = _split_table_name(target_table)
    index_columns = [*natural_key_cols, current_ind]
    index_name = f"ix_{_safe_identifier_token(table)}_{'_'.join(_safe_identifier_token(col) for col in index_columns)}"
    rendered_columns = ", ".join(f'"{col}"' for col in index_columns)
    return f"CREATE INDEX {index_name} ON {target_table} ({rendered_columns})"


def _normalize_table_key(schema: str | None, table: str) -> tuple[str | None, str]:
    normalized_schema = schema.casefold() if schema else None
    return normalized_schema, table.casefold()


def _build_dimension_table_index(
    metadata_registry: Mapping[str, DimensionMetadata],
) -> dict[tuple[str | None, str], list[str]]:
    index: dict[tuple[str | None, str], list[str]] = {}
    for name, metadata in metadata_registry.items():
        schema, table = _split_table_name(metadata.target_table)
        key = _normalize_table_key(schema, table)
        index.setdefault(key, []).append(name)
    return index


def _resolve_dimension_dependency_name(
    dependency: DependencyJoin,
    table_index: Mapping[tuple[str | None, str], list[str]],
) -> str | None:
    schema = dependency.schema_name
    table = dependency.table
    if schema is None:
        schema, table = _split_table_name(table)
    key = _normalize_table_key(schema, table)
    direct_matches = table_index.get(key, [])
    if direct_matches:
        if len(direct_matches) > 1:
            raise DimensionDependencyError(
                f"Dependency '{dependency.table}' is ambiguous across registered dimensions: {sorted(direct_matches)}"
            )
        return direct_matches[0]
    if schema is not None:
        return None

    table_matches = sorted(
        {
            name
            for (candidate_schema, candidate_table), names in table_index.items()
            if candidate_table == table.casefold()
            for name in names
        }
    )
    if not table_matches:
        return None
    if len(table_matches) > 1:
        raise DimensionDependencyError(
            f"Dependency '{dependency.table}' is ambiguous across registered dimensions: {table_matches}"
        )
    return table_matches[0]


def _find_dimension_dependency_cycle(graph: Mapping[str, set[str]]) -> list[str]:
    visited: set[str] = set()
    stack: list[str] = []
    active: set[str] = set()

    def _visit(node: str) -> list[str] | None:
        visited.add(node)
        active.add(node)
        stack.append(node)
        for dependency in sorted(graph[node]):
            if dependency not in visited:
                cycle = _visit(dependency)
                if cycle:
                    return cycle
            elif dependency in active:
                start = stack.index(dependency)
                return stack[start:] + [dependency]
        stack.pop()
        active.remove(node)
        return None

    for node in sorted(graph):
        if node not in visited:
            cycle = _visit(node)
            if cycle:
                return cycle
    return []



def _expected_column_names(metadata: DimensionMetadata) -> set[str]:
    names = {metadata.surrogate_key, metadata.join_numeric_key}
    names.update(metadata.natural_key_cols)
    names.update(metadata.data_columns.keys())
    names.update(
        value
        for key, value in DEFAULT_SYSTEM_COLUMNS.items()
        if key not in {"surrogate_key", "join_numeric_key"}
    )
    return names


def _system_column_comments(metadata: DimensionMetadata | None = None) -> dict[str, str]:
    comments = {
        DEFAULT_SYSTEM_COLUMNS["batch_id"]: "Batch identifier for this run",
        DEFAULT_SYSTEM_COLUMNS["archive_filename"]: "Archive file name for the snapshot",
        DEFAULT_SYSTEM_COLUMNS["row_hash"]: "Hash of all data columns",
        DEFAULT_SYSTEM_COLUMNS["current_ind"]: "Indicates whether row is current",
        DEFAULT_SYSTEM_COLUMNS["deleted_ind"]: "Indicates whether row is marked deleted",
        DEFAULT_SYSTEM_COLUMNS["insert_date"]: "Insert timestamp",
        DEFAULT_SYSTEM_COLUMNS["update_date"]: "Last update timestamp",
        DEFAULT_SYSTEM_COLUMNS["effective_date_start"]: "Effective start timestamp",
        DEFAULT_SYSTEM_COLUMNS["effective_date_end"]: "Effective end timestamp",
    }
    if metadata:
        comments.setdefault(
            metadata.surrogate_key, "Surrogate key, unique identifier within the table"
        )
        comments.setdefault(
            metadata.join_numeric_key, "Increasing numeric key used for joins"
        )
    return comments


def _column_comments(metadata: DimensionMetadata) -> dict[str, str]:
    comments = _system_column_comments(metadata)
    for spec_map in (metadata.natural_key_specs, metadata.data_columns):
        for name, spec in spec_map.items():
            if spec.comment:
                comments[name] = spec.comment
    return comments


def _apply_table_description(engine: Engine, metadata: DimensionMetadata) -> None:
    if engine.dialect.name != "mssql":
        return

    schema, table = _split_table_name(metadata.target_table)
    schema = schema or "dbo"
    column_comments = _column_comments(metadata)
    with engine.begin() as conn:
        _apply_table_and_column_comments(
            conn, schema, table, metadata.description, column_comments
        )


def _apply_table_and_column_comments(
    conn: Connection,
    schema: str,
    table: str,
    description: str | None,
    column_comments: dict[str, str],
) -> None:
    statements = [
        "DECLARE @schema SYSNAME = :schema;",
        "DECLARE @table SYSNAME = :table;",
    ]
    params = {"schema": schema, "table": table}
    if description:
        safe_desc = description.replace("'", "''")
        statements.append(
            "DECLARE @description NVARCHAR(4000) = :description;"
        )
        statements.append(
            f"-- description: {safe_desc}"
        )
        params["description"] = description
    statements.append(
        "DECLARE @columns TABLE (ColumnName SYSNAME, Comment NVARCHAR(4000));"
    )
    for idx, (column, comment) in enumerate(column_comments.items()):
        key_col = f"column_{idx}"
        key_comment = f"comment_{idx}"
        statements.append(f"-- column: {column}")
        statements.append(
            f"INSERT INTO @columns (ColumnName, Comment) VALUES (:{key_col}, :{key_comment});"
        )
        params[key_col] = column
        params[key_comment] = comment
    if description:
        statements.append(
            """
BEGIN TRY
  EXEC sys.sp_updateextendedproperty
    @name=N'MS_Description', @value=@description,
    @level0type=N'SCHEMA', @level0name=@schema,
    @level1type=N'TABLE', @level1name=@table;
END TRY
BEGIN CATCH
  EXEC sys.sp_addextendedproperty
    @name=N'MS_Description', @value=@description,
    @level0type=N'SCHEMA', @level0name=@schema,
    @level1type=N'TABLE', @level1name=@table;
END CATCH;
"""
        )
    statements.append(
        """
DECLARE @col SYSNAME, @comment NVARCHAR(4000);
WHILE EXISTS (SELECT 1 FROM @columns)
BEGIN
  SELECT TOP 1 @col = ColumnName, @comment = Comment FROM @columns;
  BEGIN TRY
    EXEC sys.sp_updateextendedproperty
      @name=N'MS_Description', @value=@comment,
      @level0type=N'SCHEMA', @level0name=@schema,
      @level1type=N'TABLE', @level1name=@table,
      @level2type=N'COLUMN', @level2name=@col;
  END TRY
      BEGIN CATCH
    EXEC sys.sp_addextendedproperty
      @name=N'MS_Description', @value=@comment,
      @level0type=N'SCHEMA', @level0name=@schema,
      @level1type=N'TABLE', @level1name=@table,
      @level2type=N'COLUMN', @level2name=@col;
  END CATCH;
  DELETE FROM @columns WHERE ColumnName = @col;
END
"""
    )
    script = "\n".join(statements)
    conn.execute(text(script), params)
    logger.info("Applied comments to %s.%s", schema, table)


def _apply_system_column_comments(
    conn: Connection,
    schema: str,
    table: str,
    column_comments: dict[str, str],
) -> None:
    """Apply column comments without a table-level description."""
    _apply_table_and_column_comments(conn, schema, table, None, column_comments)

def _column_spec_for_missing(name: str, metadata: DimensionMetadata) -> ColumnSpec | None:
    if name in metadata.natural_key_specs:
        return metadata.natural_key_specs[name]
    if name in metadata.natural_key_cols:
        return ColumnSpec(type=None, nullable=False)
    if name in {
        DEFAULT_SYSTEM_COLUMNS["effective_date_start"],
        DEFAULT_SYSTEM_COLUMNS["effective_date_end"],
    }:
        return ColumnSpec(type="DATETIME2(3)", nullable=False)
    return metadata.data_columns.get(name)


def _evolve_table_columns(
    engine: Engine,
    metadata: DimensionMetadata,
    missing_columns: set[str],
) -> None:
    with engine.begin() as conn:
        for column in sorted(missing_columns):
            if column in {
                DEFAULT_SYSTEM_COLUMNS["effective_date_start"],
                DEFAULT_SYSTEM_COLUMNS["effective_date_end"],
            }:
                _evolve_effective_date_column(conn, engine, metadata.target_table, column)
                logger.info(
                    "Evolved table %s by adding column %s for mode=%s",
                    metadata.target_table,
                    column,
                    metadata.schema_handling.mode,
                )
                continue
            spec = _column_spec_for_missing(column, metadata)
            if spec is None:
                logger.warning("Cannot evolve column '%s': no spec defined", column)
                continue
            clause = _column_spec_clause(column, spec, engine)
            add_sql = text(f"ALTER TABLE {metadata.target_table} ADD {clause}")
            conn.execute(add_sql)
            logger.info("Evolved table %s by adding column %s for mode=%s", metadata.target_table, column, metadata.schema_handling.mode)


def _evolve_effective_date_column(conn: Connection, engine: Engine, table_name: str, column: str) -> None:
    column_type = _system_column_type(column, engine)
    quoted_column = _quote_identifier(engine, column)
    insert_date_column = _quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS["insert_date"])
    current_ind_column = _quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS["current_ind"])
    deleted_ind_column = _quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS["deleted_ind"])
    update_date_column = _quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS["update_date"])
    conn.execute(text(f"ALTER TABLE {table_name} ADD {quoted_column} {column_type}"))
    if column == DEFAULT_SYSTEM_COLUMNS["effective_date_start"]:
        conn.execute(
            text(
                f"""
                UPDATE {table_name}
                SET {quoted_column} = {insert_date_column}
                WHERE {quoted_column} IS NULL
                """
            )
        )
    else:
        conn.execute(
            text(
                f"""
                UPDATE {table_name}
                SET {quoted_column} = CASE
                    WHEN {current_ind_column} = 1
                         AND {deleted_ind_column} = 0
                    THEN :effective_date_max
                    ELSE COALESCE({update_date_column}, :effective_date_max)
                END
                WHERE {quoted_column} IS NULL
                """
            ),
            {"effective_date_max": EFFECTIVE_DATE_MAX},
        )
    if engine.dialect.name == "mssql":
        conn.execute(text(f"ALTER TABLE {table_name} ALTER COLUMN {quoted_column} {column_type} NOT NULL"))

def _surrogate_column_clause(engine: Engine, name: str) -> str:
    dialect = engine.dialect.name
    if dialect == "mssql":
        return f"{_quote_identifier(engine, name)} INT IDENTITY(1,1) PRIMARY KEY"
    return f"{_quote_identifier(engine, name)} INTEGER PRIMARY KEY AUTOINCREMENT"


def _join_numeric_clause(engine: Engine, name: str) -> str:
    dialect = engine.dialect.name
    if dialect == "mssql":
        return f"{_quote_identifier(engine, name)} INT NOT NULL"
    return f"{_quote_identifier(engine, name)} INTEGER NOT NULL"


def _resolve_safe_data_columns(
    requested: Sequence[str], available: set[str]
) -> tuple[list[str], list[str]]:
    safe = [col for col in requested if col in available]
    missing = [col for col in requested if col not in available]
    if not safe and missing:
        raise RuntimeError("No safe data columns available to write.")
    return safe, missing


def _apply_dependency_joins(
    incoming: pd.DataFrame,
    dependencies: Sequence[DependencyJoin],
    engine: Engine,
) -> pd.DataFrame:
    for dep in dependencies:
        schema, table = dep.schema_name, dep.table
        if schema is None:
            schema, table = _split_table_name(table)
        table_ref = f"[{schema}].[{table}]" if schema else f"[{table}]"
        on_source = [relation["source"] for relation in dep.on]
        on_target = [relation["target"] for relation in dep.on]
        select_aliases = dict(dep.select)
        select_columns = on_target + list(select_aliases.keys())
        quoted_cols = ", ".join(f"[{col}]" for col in select_columns)
        where_clauses = []
        if dep.filter_target_current:
            where_clauses.append("[Current_Ind] = 1")
            try:
                inspector = inspect(engine)
                dep_columns = {
                    column["name"] for column in inspector.get_columns(table, schema=schema)
                }
            except Exception:
                dep_columns = set()
            if "Deleted_Ind" in dep_columns:
                where_clauses.append("[Deleted_Ind] = 0")
        for expressions in dep.where.values():
            where_clauses.extend(expressions)
        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        logger.debug(
            "Running dependency join against %s (filter_current=%s) for %d rows; natural keys=%s",
            table_ref,
            dep.filter_target_current,
            len(incoming),
            incoming[on_source].drop_duplicates().head(5).to_dict(orient="records"),
        )

        query = text(f"SELECT {quoted_cols} FROM {table_ref} {where_sql}")
        dep_df = pd.read_sql_query(query, con=engine)

        dep_df = dep_df.rename(columns=dict(zip(on_target, on_source)))
        dep_df = dep_df.rename(columns=select_aliases)

        duplicate_keys = dep_df.duplicated(subset=on_source, keep=False)
        if duplicate_keys.any():
            samples = dep_df.loc[duplicate_keys, on_source].drop_duplicates().head(5)
            raise RuntimeError(
                f"Dependency join {dep.table} produced duplicate matches for keys "
                f"{samples.to_dict(orient='records')}"
            )

        dep_map = {}
        for _, row in dep_df.iterrows():
            key = tuple(_normalize_dependency_join_key(row[src]) for src in on_source)
            dep_map[key] = {alias: row[alias] for alias in select_aliases.values()}

        key_series = incoming[on_source].apply(
            lambda row: tuple(_normalize_dependency_join_key(row[src]) for src in on_source),
            axis=1,
        )
        matched_keys = key_series.map(lambda key: key in dep_map)
        for alias in select_aliases.values():
            incoming[alias] = key_series.map(lambda key: dep_map.get(key, {}).get(alias))

        if dep.how == "inner":
            incoming = incoming.loc[matched_keys].copy()
            matched_keys = matched_keys.loc[incoming.index]

        missing_mask = ~matched_keys
        if missing_mask.any():
            missing_rows = incoming.loc[missing_mask, on_source + list(select_aliases.values())]
            logger.info(
                "Dependency join %s produced %s missing rows; sample outputs: %s",
                table_ref,
                missing_mask.sum(),
                missing_rows.head(5).to_dict(orient="records"),
            )
            if dep.on_missing == "error":
                raise RuntimeError(
                    _format_missing_dependency_join_error(
                        dep=dep,
                        table_ref=table_ref,
                        missing_rows=missing_rows,
                    )
                )

    return incoming


def _normalize_dependency_join_key(value: Any) -> str | None:
    if pd.isna(value):
        return None
    value = _normalize_value_for_sql(value)
    if isinstance(value, bytes):
        value = value.decode("utf-8", errors="replace")
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _format_missing_dependency_join_error(
    *,
    dep: DependencyJoin,
    table_ref: str,
    missing_rows: pd.DataFrame,
) -> str:
    example_missing_keys = missing_rows[[relation["source"] for relation in dep.on]]
    examples = example_missing_keys.drop_duplicates().head(10).to_dict(orient="records")
    return (
        f"Missing dependency match in {table_ref} for source columns "
        f"{[relation['source'] for relation in dep.on]} -> target columns "
        f"{[relation['target'] for relation in dep.on]}. "
        f"Required columns {list(dep.select.values())} were null for {len(missing_rows)} row(s). "
        f"Example missing source keys: {examples}"
    )


def _raise_incompatible_type_error(series: pd.Series, invalid_mask: pd.Series, column: str, target_type: str) -> None:
    invalid_values = series.loc[invalid_mask].drop_duplicates().head(3).tolist()
    raise ValueError(
        f"Column '{column}' contains value(s) incompatible with target type {target_type}: "
        f"{invalid_values}. SCD2 load aborted."
    )


def _coerce_boolean_series(series: pd.Series, column: str, target_type: str) -> pd.Series:
    true_values = {"1", "true", "t", "yes", "y"}
    false_values = {"0", "false", "f", "no", "n"}
    coerced = []
    invalid_mask = pd.Series(False, index=series.index, dtype="boolean")
    for index, value in series.items():
        if pd.isna(value):
            coerced.append(pd.NA)
            continue
        if isinstance(value, bool):
            coerced.append(value)
            continue
        if hasattr(value, "item"):
            scalar_value = value.item()
            if isinstance(scalar_value, bool):
                coerced.append(scalar_value)
                continue
        if isinstance(value, int) and value in {0, 1}:
            coerced.append(bool(value))
            continue
        if isinstance(value, float) and value in {0.0, 1.0}:
            coerced.append(bool(int(value)))
            continue
        if isinstance(value, str):
            normalized = value.strip().lower()
            if normalized in true_values:
                coerced.append(True)
                continue
            if normalized in false_values:
                coerced.append(False)
                continue
        coerced.append(pd.NA)
        invalid_mask.loc[index] = True
    if invalid_mask.any():
        _raise_incompatible_type_error(series, invalid_mask, column, target_type)
    return pd.Series(coerced, index=series.index, dtype="boolean")


def _coerce_uuid_series(series: pd.Series, column: str, target_type: str) -> pd.Series:
    coerced = []
    invalid_mask = pd.Series(False, index=series.index, dtype="boolean")
    for index, value in series.items():
        if pd.isna(value):
            coerced.append(None)
            continue
        try:
            coerced.append(str(uuid.UUID(str(value).strip())))
        except (AttributeError, TypeError, ValueError):
            coerced.append(None)
            invalid_mask.loc[index] = True
    if invalid_mask.any():
        _raise_incompatible_type_error(series, invalid_mask, column, target_type)
    return pd.Series(coerced, index=series.index, dtype="object")


def _format_temporal_value_for_sql(value: datetime, base_sql_type: str) -> str:
    if base_sql_type == "DATE":
        return value.date().isoformat()
    if base_sql_type == "TIME":
        return value.time().isoformat(timespec="milliseconds")
    return value.isoformat(timespec="milliseconds")


def _parse_temporal_value_for_sql(
    value: object,
    *,
    fmt: str,
    base_sql_type: str,
) -> str | None:
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return _format_temporal_value_for_sql(value, base_sql_type)
    if not isinstance(value, str):
        return None
    raw = value.strip()
    parse_attempts = [
        lambda text: datetime.strptime(text, fmt),
        lambda text: datetime.fromisoformat(text.replace("Z", "+00:00")),
    ]
    for parser in parse_attempts:
        try:
            return _format_temporal_value_for_sql(parser(raw), base_sql_type)
        except ValueError:
            continue
    return None


def _normalize_parsed_temporal_series_for_sql(parsed: pd.Series, base_sql_type: str) -> pd.Series:
    return parsed.map(
        lambda value: None
        if pd.isna(value)
        else _format_temporal_value_for_sql(value.to_pydatetime(), base_sql_type)
    )


def _cast_data_columns(incoming: pd.DataFrame, metadata: DimensionMetadata) -> pd.DataFrame:
    integer_types = {"INT", "INTEGER", "BIGINT", "SMALLINT", "TINYINT"}
    decimal_types = {"FLOAT", "REAL", "DECIMAL", "NUMERIC"}
    boolean_types = {"BIT", "BOOLEAN", "BOOL"}
    uuid_types = {"UNIQUEIDENTIFIER", "UUID"}
    df = incoming.copy()
    for name, spec in metadata.data_columns.items():
        if name not in df.columns:
            continue
        if not spec.type:
            continue
        sql_type = spec.type.upper()
        base_sql_type = re.split(r"[\s(]", sql_type, maxsplit=1)[0]
        if "DATETIME" in sql_type or base_sql_type in {"DATE", "TIME"}:
            fmt = spec.parse_format or DEFAULT_DATETIME_INPUT_FORMAT
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore",
                    message="Could not infer format, so each element will be parsed individually.*",
                )
                parsed = pd.to_datetime(
                    df[name],
                    format=fmt,
                    errors="coerce",
                    dayfirst=spec.dayfirst,
                )
                retry_mask = df[name].notna() & parsed.isna()
                if retry_mask.any():
                    parsed.loc[retry_mask] = pd.to_datetime(
                        df.loc[retry_mask, name],
                        errors="coerce",
                        dayfirst=spec.dayfirst,
                    )
                invalid_mask = df[name].notna() & parsed.isna()
                if invalid_mask.any():
                    normalized = {}
                    for index, value in df.loc[invalid_mask, name].items():
                        normalized_value = _parse_temporal_value_for_sql(
                            value,
                            fmt=fmt,
                            base_sql_type=base_sql_type,
                        )
                        if normalized_value is not None:
                            normalized[index] = normalized_value
                    unresolved_mask = invalid_mask.copy()
                    for index in normalized:
                        unresolved_mask.loc[index] = False
                    if unresolved_mask.any():
                        _raise_incompatible_type_error(df[name], unresolved_mask, name, spec.type)
                    normalized_series = _normalize_parsed_temporal_series_for_sql(parsed, base_sql_type)
                    for index, value in normalized.items():
                        normalized_series.loc[index] = value
                    df[name] = normalized_series.astype("object")
                    continue
                if base_sql_type == "TIME":
                    df[name] = parsed.dt.time
                else:
                    df[name] = parsed
            continue
        if base_sql_type in decimal_types | integer_types:
            parsed = pd.to_numeric(df[name], errors="coerce")
            invalid_mask = df[name].notna() & parsed.isna()
            if base_sql_type in integer_types:
                non_null_mask = parsed.notna()
                whole_number_mask = (parsed[non_null_mask] % 1).eq(0)
                invalid_mask.loc[non_null_mask] = invalid_mask.loc[non_null_mask] | ~whole_number_mask
            if invalid_mask.any():
                _raise_incompatible_type_error(df[name], invalid_mask, name, spec.type)
            if base_sql_type in integer_types:
                df[name] = parsed.astype("Int64")
            else:
                df[name] = parsed
            continue
        if base_sql_type in boolean_types:
            df[name] = _coerce_boolean_series(df[name], name, spec.type)
            continue
        if base_sql_type in uuid_types:
            df[name] = _coerce_uuid_series(df[name], name, spec.type)
    return df


def _handle_schema_drift(
    engine: Engine,
    metadata: DimensionMetadata,
    available_columns: set[str],
    requested_columns: Mapping[str, ColumnSpec],
) -> tuple[list[str], list[str]]:
    requested_names = list(requested_columns.keys())
    safe, missing = _resolve_safe_data_columns(requested_names, available_columns)
    if missing:
        plan = f"Missing columns: {missing}"
        if metadata.schema_handling.mode == "evolve":
            missing_specs = {name: requested_columns[name] for name in missing}
            _evolve_schema(engine, metadata.target_table, missing_specs)
            safe = list(requested_columns)
            logger.info("Schema evolution applied for %s: added %s", metadata.target_table, missing)
        else:
            logger.warning(
                "Schema evolution plan for %s: %s",
                metadata.target_table,
                plan,
            )
    return safe, missing


def _column_type_for_engine(engine: Engine) -> str:
    dialect = engine.dialect.name
    if dialect == "mssql":
        return "NVARCHAR(4000)"
    return "TEXT"


def _system_column_type(column: str, engine: Engine) -> str:
    dialect = engine.dialect.name
    if column in {DEFAULT_SYSTEM_COLUMNS["current_ind"], DEFAULT_SYSTEM_COLUMNS["deleted_ind"]}:
        if dialect == "mssql":
            return "BIT"
        return "BOOLEAN"
    if column == DEFAULT_SYSTEM_COLUMNS["row_hash"]:
        if dialect == "mssql":
            return "NVARCHAR(4000)"
        return "TEXT"
    if column in {DEFAULT_SYSTEM_COLUMNS["batch_id"], DEFAULT_SYSTEM_COLUMNS["archive_filename"]}:
        if dialect == "mssql":
            return "NVARCHAR(4000)"
        return "TEXT"
    if dialect == "mssql":
        return "DATETIME2(3)"
    return "DATETIME"


def _evolve_schema(engine: Engine, table_name: str, missing: Mapping[str, ColumnSpec]) -> None:
    with engine.begin() as conn:
        for column, spec in missing.items():
            column_type = spec.type or _column_type_for_engine(engine)
            constraints = []
            if not spec.nullable:
                constraints.append("NOT NULL")
            if spec.unique:
                constraints.append("UNIQUE")
            if spec.default is not None:
                constraints.append(f"DEFAULT {spec.default}")
            constraint_sql = " ".join(constraints)
            definition = f"{column_type} {constraint_sql}".strip()
            conn.execute(
                text(f"ALTER TABLE {table_name} ADD COLUMN {_quote_identifier(engine, column)} {definition}")
            )


def _capture_execution_time() -> datetime:
    now = datetime.now(timezone.utc)
    milliseconds = (now.microsecond // 1000) * 1000
    return now.replace(microsecond=milliseconds)


def _post_scd2_validation(engine: Engine, table: str, natural_key_cols: Sequence[str]) -> None:
    column_types = _get_table_column_types(engine, table)
    group_expr = ", ".join(
        _case_sensitive_group_expression(engine, col, column_types)
        for col in natural_key_cols
    )
    key_expr = ", ".join(
        f"{_case_sensitive_group_expression(engine, col, column_types)} AS {_quote_identifier(engine, col)}"
        for col in natural_key_cols
    )
    current_ind_column = _quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS["current_ind"])
    deleted_ind_column = _quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS["deleted_ind"])
    update_date_column = _quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS["update_date"])
    effective_end_column = _quote_identifier(engine, DEFAULT_SYSTEM_COLUMNS["effective_date_end"])
    current_dups_sql = text(
        f"""
        SELECT {key_expr}, COUNT(*) AS cnt
        FROM {table}
        WHERE {current_ind_column} = 1
        GROUP BY {group_expr}
        HAVING COUNT(*) > 1
        """
    )

    dialect = engine.dialect.name
    top = "TOP 1 " if dialect == "mssql" else ""
    limit = "" if dialect == "mssql" else " LIMIT 1"
    validation_checks = [
        (
            text(
                f"SELECT {top}1 FROM {table} WHERE {current_ind_column} = 1 AND {deleted_ind_column} = 0 AND {update_date_column} IS NOT NULL{limit}"
            ),
            "current row has Update_Date not NULL",
        ),
        (
            text(
                f"SELECT {top}1 FROM {table} WHERE {current_ind_column} = 1 AND {deleted_ind_column} = 0 AND {effective_end_column} != :effective_date_max{limit}"
            ),
            "current row has Effective_Date_End not at max",
        ),
        (
            text(
                f"SELECT {top}1 FROM {table} WHERE {current_ind_column} = 0 AND {update_date_column} IS NULL{limit}"
            ),
            "historical row missing Update_Date",
        ),
    ]

    with engine.connect() as conn:
        duplicate_rows = conn.execute(current_dups_sql).mappings().fetchmany(10)
        if duplicate_rows:
            examples = [
                {column: row[column] for column in natural_key_cols}
                for row in duplicate_rows
            ]
            raise RuntimeError(
                f"Multiple current rows found for natural key columns {list(natural_key_cols)}. "
                f"Example duplicate keys: {examples}"
            )
        for sql_stmt, message in validation_checks:
            if conn.execute(sql_stmt, {"effective_date_max": EFFECTIVE_DATE_MAX}).first():
                raise RuntimeError(f"Post-SCD2 validation failed: {message}")


def _get_table_column_types(engine: Engine, table_name: str) -> dict[str, str]:
    inspector = inspect(engine)
    schema, table = _split_table_name(table_name)
    return {
        column["name"].lower(): str(column["type"])
        for column in inspector.get_columns(table, schema=schema)
    }


def _is_text_like_type(type_name: str | None) -> bool:
    if not type_name:
        return False
    normalized = type_name.strip().upper()
    return any(token in normalized for token in ("CHAR", "TEXT", "CLOB"))


def _case_sensitive_group_expression(
    engine: Engine,
    column: str,
    column_types: Mapping[str, str],
) -> str:
    quoted = _quote_identifier(engine, column)
    if not _is_text_like_type(column_types.get(column.lower())):
        return quoted
    if engine.dialect.name == "sqlite":
        return f"{quoted} COLLATE BINARY"
    if engine.dialect.name == "mssql":
        return f"{quoted} COLLATE Latin1_General_100_BIN2"
    return quoted


def _resolve_archive_destination(
    metadata: DimensionMetadata, execution_time: datetime
) -> Path | None:
    if not metadata.archive_path:
        return None

    base = Path(metadata.archive_path)
    if base.suffix:
        return base

    suffix = ".parquet"
    return base / f"{metadata.name}_{execution_time.strftime('%Y%m%d%H%M%S')}{suffix}"


def _archive_snapshot(
    incoming: pd.DataFrame,
    metadata: DimensionMetadata,
    archive_path: Path | None,
) -> None:
    if not archive_path:
        return

    archive_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        _prepare_archive_snapshot(incoming).to_parquet(archive_path, index=False)
        logger.info("Archived snapshot to %s", archive_path)
    except Exception as exc:
        logger.warning("Failed to archive snapshot to %s: %s", archive_path, exc)


def _prepare_archive_snapshot(incoming: pd.DataFrame) -> pd.DataFrame:
    archive_df = incoming.copy()
    for column in archive_df.columns:
        series = archive_df[column]
        if series.dtype != "object":
            continue
        non_null = series.dropna()
        if non_null.empty:
            continue
        value_types = {type(value) for value in non_null}
        if len(value_types) <= 1:
            continue
        archive_df[column] = series.map(
            lambda value: None if pd.isna(value) else str(value)
        )
    return archive_df
