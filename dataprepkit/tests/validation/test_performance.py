from time import perf_counter

import openpyxl

from dataprepkit.validation import (
    ColumnDefinition,
    DataBoundary,
    DataFrameLoadPolicy,
    HeaderPolicy,
    RuntimePolicy,
    SheetSelector,
    TableConfig,
    validate_excel,
)

from test_public_api import make_config


def test_large_workbook_validation_timing(tmp_path):
    path = tmp_path / "large.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    for row_number in range(1, 5001):
        sheet.cell(row_number, 1).value = row_number
        sheet.cell(row_number, 2).value = f"value-{row_number}"
    workbook.save(path)

    started = perf_counter()
    result = validate_excel(
        candidate_path=path,
        config=make_config().model_copy(
            update={
                "runtime": RuntimePolicy(
                    max_cells_scanned=20_000,
                    read_only=False,
                    macro_policy="reject",
                    missing_formula_cache_action="not_run",
                )
            }
        ),
    )
    elapsed = perf_counter() - started

    assert result.is_valid is True
    assert elapsed < 10


def test_read_only_inflated_dimension_does_not_scan_empty_rows(tmp_path):
    path = tmp_path / "inflated.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1048576"] = "last-row-value"
    workbook.save(path)

    config = make_config().model_copy(
        update={
            "runtime": RuntimePolicy(
                max_cells_scanned=10,
                read_only=True,
                macro_policy="reject",
                missing_formula_cache_action="not_run",
            )
        }
    )
    started = perf_counter()
    result = validate_excel(candidate_path=path, config=config)
    elapsed = perf_counter() - started

    assert result.is_valid is True
    assert result.errors == []
    assert elapsed < 3


def test_large_pandas_table_load_timing(tmp_path):
    path = tmp_path / "large_table.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Measure_Value"])
    for row_number in range(1, 5001):
        sheet.append([f"value-{row_number}"])
    workbook.save(path)

    table = TableConfig(
        name="data",
        sheet_selector=SheetSelector(mode="exact", value="Data"),
        header_row=1,
        column_definitions=[ColumnDefinition(name="Measure_Value")],
        header_policy=HeaderPolicy(required_columns=["Measure_Value"]),
        data_boundary=DataBoundary(
            mode="last_non_empty_row",
            columns=["Measure_Value"],
        ),
        load_policy=DataFrameLoadPolicy(),
    )
    config = make_config().model_copy(
        update={
            "tables": [table],
            "runtime": RuntimePolicy(
                max_cells_scanned=20_000,
                read_only=False,
                macro_policy="reject",
                missing_formula_cache_action="not_run",
            ),
        }
    )

    started = perf_counter()
    result = validate_excel(candidate_path=path, config=config)
    elapsed = perf_counter() - started

    assert result.is_valid is True
    assert result.errors == []
    assert elapsed < 10
