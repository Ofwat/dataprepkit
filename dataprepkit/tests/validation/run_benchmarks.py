"""Run the Excel validation benchmarks and emit GitHub benchmark JSON."""

import argparse
import json
import sys
from pathlib import Path
from tempfile import TemporaryDirectory
from time import perf_counter

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

from dataprepkit.validation import RuntimePolicy, validate_excel

from test_public_api import make_config


def _run(path, config):
    started = perf_counter()
    result = validate_excel(candidate_path=path, config=config)
    elapsed = perf_counter() - started
    if not result.is_valid:
        raise RuntimeError(result.format_report())
    return elapsed


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    with TemporaryDirectory() as directory:
        root = Path(directory)
        large_path = root / "large.xlsx"
        large_workbook = openpyxl.Workbook()
        large_sheet = large_workbook.active
        large_sheet.title = "Data"
        for row_number in range(1, 5001):
            large_sheet.cell(row_number, 1).value = row_number
            large_sheet.cell(row_number, 2).value = f"value-{row_number}"
        large_workbook.save(large_path)

        inflated_path = root / "inflated.xlsx"
        inflated_workbook = openpyxl.Workbook()
        inflated_workbook.active.title = "Data"
        inflated_workbook.active["A1048576"] = "last-row-value"
        inflated_workbook.save(inflated_path)

        large_config = make_config().model_copy(
            update={
                "runtime": RuntimePolicy(
                    max_cells_scanned=20_000,
                    read_only=False,
                    macro_policy="reject",
                    missing_formula_cache_action="not_run",
                )
            }
        )
        inflated_config = make_config().model_copy(
            update={
                "runtime": RuntimePolicy(
                    max_cells_scanned=10,
                    read_only=True,
                    macro_policy="reject",
                    missing_formula_cache_action="not_run",
                )
            }
        )
        benchmarks = [
            {
                "name": "excel_validation_5000_rows_2_columns",
                "unit": "seconds",
                "value": _run(large_path, large_config),
            },
            {
                "name": "excel_validation_read_only_inflated_A1048576",
                "unit": "seconds",
                "value": _run(inflated_path, inflated_config),
            },
        ]

    args.output.write_text(json.dumps(benchmarks, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
