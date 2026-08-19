from datetime import datetime
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.formula import ArrayFormula


from dataprepkit.validation import (
    ComparisonConfig,
    ConfigurationError,
    CellComparison,
    CellLocation,
    ColumnDefinition,
    ColumnValidation,
    DataBoundary,
    EmptyRowRule,
    ExpectedCellCheck,
    HeaderPolicy,
    RuntimePolicy,
    WorkbookFeaturePolicy,
    SheetPolicy,
    SheetSelector,
    TableConfig,
    ValidationEvent,
    ValidationResult,
    WorkbookValidationConfig,
    WorkbookCheck,
    dump_validation_config,
    dump_validation_result,
    get_config_schema,
    list_profiles,
    list_registered_rules,
    register_rule,
    unregister_rule,
    load_validation_config,
    load_validation_result,
    validate_config,
    validate_excel,
)


def make_config(
    required_sheet="Data",
    workbook_checks=None,
    extra_sheet_action="ignore",
    ignored_selectors=None,
):
    return WorkbookValidationConfig(
        config_version="1",
        comparison=ComparisonConfig(
            case_sensitive=False,
            accent_sensitive=True,
            trim_whitespace=True,
            collapse_internal_whitespace=False,
            empty_string_is_null=True,
            null_tokens=[],
            collation_name="test-collation",
        ),
        sheet_policy=SheetPolicy(
            required_selectors=[
                SheetSelector(mode="exact", value=required_sheet)
            ],
            ignored_selectors=ignored_selectors or [],
            extra_sheet_action=extra_sheet_action,
            selector_match_action="all",
        ),
        tables=[],
        expected_cells=[],
        workbook_checks=workbook_checks or [],
        enabled_rules=None,
        rule_severity={},
        runtime=RuntimePolicy(
            max_cells_scanned=1000,
            read_only=False,
            macro_policy="reject",
            missing_formula_cache_action="not_run",
        ),
        compatibility=None,
    )


def write_workbook(path: Path, sheet_name: str):
    workbook = openpyxl.Workbook()
    workbook.active.title = sheet_name
    workbook.save(path)


def write_workbook_with_sheets(path: Path, sheet_names):
    workbook = openpyxl.Workbook()
    workbook.active.title = sheet_names[0]
    for sheet_name in sheet_names[1:]:
        workbook.create_sheet(sheet_name)
    workbook.save(path)


def test_public_validation_exports_are_available():
    assert WorkbookValidationConfig
    assert validate_config
    assert validate_excel


def test_custom_rule_registration_runs_through_public_validator(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")

    rule_code = "custom_test_rule"

    def validator(context):
        return ValidationEvent(
            rule_code=context.rule_code,
            sheet_name="Data",
            description="Custom rule failed",
        )

    register_rule(rule_code, validator)
    try:
        assert rule_code in list_registered_rules()
        config = make_config(
            workbook_checks=[
                WorkbookCheck(
                    rule_code=rule_code,
                    enabled=True,
                    scope="all_sheets",
                )
            ]
        )
        result = validate_excel(
            candidate_path=candidate_path,
            config=config,
        )
    finally:
        unregister_rule(rule_code)

    assert [event.rule_code for event in result.errors] == [rule_code]


def test_validate_config_returns_a_validated_public_model():
    config = make_config()

    resolved = validate_config(config)

    assert resolved is config
    assert resolved.config_version == "1"


def test_validation_config_round_trips_through_mapping_and_json():
    config = make_config()

    mapping = dump_validation_config(config, format="mapping")
    assert load_validation_config(mapping) == config

    encoded = dump_validation_config(config, format="json")
    assert load_validation_config(encoded) == config


def test_get_config_schema_returns_versioned_public_schema():
    schema = get_config_schema()

    assert schema["$id"].endswith("/validation-config/v1")
    assert schema["title"] == "WorkbookValidationConfig"
    assert schema["type"] == "object"
    assert schema["additionalProperties"] is False


def test_get_config_schema_rejects_unsupported_version():
    with pytest.raises(ConfigurationError):
        get_config_schema(version="2")


def test_list_profiles_returns_deterministic_profile_metadata(tmp_path):
    profile_path = tmp_path / "baseline.yaml"
    profile_path.write_text(
        dump_validation_config(
            make_config().model_copy(
                update={
                    "profile_name": "baseline",
                }
            ),
            format="yaml",
        ),
        encoding="utf-8",
    )

    assert list_profiles(tmp_path) == [
        {
            "name": "baseline",
            "config_version": "1",
            "path": str(profile_path),
        }
    ]


def test_validation_result_round_trips_through_mapping_and_json():
    result = ValidationResult(
        run_id="run-1",
        config_version="1",
        profile_name="profile",
        candidate_filename="candidate.xlsx",
    )

    mapping = dump_validation_result(result, format="mapping")
    assert load_validation_result(mapping) == result

    encoded = dump_validation_result(result, format="json")
    assert load_validation_result(encoded) == result


def test_validate_excel_populates_audit_metadata(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")
    config = make_config().model_copy(
        update={"profile_name": "profile"},
    )

    result = validate_excel(
        candidate_path=candidate_path,
        config=config,
        candidate_version="candidate-v1",
        run_id="run-1",
    )

    assert result.run_id == "run-1"
    assert result.config_version == "1"
    assert result.profile_name == "profile"
    assert result.candidate_filename == "candidate.xlsx"
    assert result.candidate_version == "candidate-v1"
    assert result.comparison == config.comparison


def test_validation_config_round_trips_through_yaml():
    config = make_config()

    encoded = dump_validation_config(config, format="yaml")

    assert load_validation_config(encoded) == config


def test_validation_config_merges_caller_overrides_over_profile():
    profile = make_config(
        extra_sheet_action="error",
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_error",
                enabled=True,
                scope="all_sheets",
                options={"error_tokens": ["#DIV/0!"]},
            )
        ],
    )
    overrides = {
        "sheet_policy": {
            "extra_sheet_action": "warning",
        },
        "workbook_checks": [],
    }

    resolved = load_validation_config(
        overrides,
        compatibility_profile=profile,
    )

    assert resolved.sheet_policy.extra_sheet_action == "warning"
    assert resolved.workbook_checks == []
    assert resolved.sheet_policy.required_selectors == (
        profile.sheet_policy.required_selectors
    )


def test_validate_config_reports_a_public_field_path():
    with pytest.raises(ConfigurationError) as error:
        validate_config(
            {
                **make_config().model_dump(),
                "tables": [
                    {
                        "name": "outputs",
                        "column_validations": [
                            {
                                "column": "status",
                                "allowed_values": ["active"],
                                "forbidden_values": ["deleted"],
                            }
                        ],
                    }
                ],
            }
        )

    assert "tables[0].column_validations[0]" in str(error.value)


def test_validate_config_rejects_incomplete_data_boundary():
    config_data = make_config().model_dump()
    config_data["tables"] = [
        {
            "name": "outputs",
            "data_boundary": {
                "mode": "fixed_end_row",
            },
        }
    ]

    with pytest.raises(ConfigurationError) as error:
        validate_config(config_data)

    assert "tables[0].data_boundary" in str(error.value)


def test_validate_config_rejects_invalid_table_policy():
    config_data = make_config().model_dump()
    config_data["tables"] = [
        {
            "name": "outputs",
            "header_policy": {
                "required_columns": ["unit"],
                "extra_column_action": "reject",
            },
        }
    ]

    with pytest.raises(ConfigurationError) as error:
        validate_config(config_data)

    assert "tables[0].header_policy" in str(error.value)


def test_validate_config_rejects_invalid_runtime_policy():
    config_data = make_config().model_dump()
    config_data["runtime"]["macro_policy"] = "calculate"

    with pytest.raises(ConfigurationError) as error:
        validate_config(config_data)

    assert "runtime" in str(error.value)

    config_data = make_config().model_dump()
    config_data["runtime"]["missing_formula_cache_action"] = "ignore"

    with pytest.raises(ConfigurationError) as error:
        validate_config(config_data)

    assert "runtime" in str(error.value)


def test_workbook_check_validates_rule_options():
    with pytest.raises(ValueError):
        WorkbookCheck(
            rule_code="formula_error",
            enabled=True,
            scope="all_sheets",
            options={"error_tokens": "#DIV/0!"},
        )

    with pytest.raises(ValueError):
        WorkbookCheck(
            rule_code="formula_difference",
            enabled=True,
            scope="overlapping_sheets",
            options={"whitespace_policy": "semantic"},
        )

    with pytest.raises(ValueError):
        WorkbookCheck(
            rule_code="formula_difference",
            enabled=True,
            scope="overlapping_sheets",
            options={"unsupported": True},
        )

    with pytest.raises(ValueError):
        ExpectedCellCheck(
            name="status",
            locations=[
                CellLocation(
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    cell_reference="A1",
                )
            ],
            expected_value="active",
            fallback_strategy="first_match",
        )


def test_expected_cell_numeric_comparison_uses_configured_separators(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = "1,000.50"
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "expected_cells": [
                ExpectedCellCheck(
                    name="amount",
                    locations=[
                        CellLocation(
                            sheet_selector=SheetSelector(
                                mode="exact",
                                value="Data",
                            ),
                            cell_reference="A1",
                        )
                    ],
                    expected_value=1000.5,
                    comparison=CellComparison(
                        mode="numeric",
                        options={
                            "decimal_separator": ".",
                            "thousands_separator": ",",
                            "allow_numeric_strings": True,
                        },
                    ),
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True


def test_expected_cell_date_comparison_uses_configured_formats(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = datetime(2025, 1, 2)
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "expected_cells": [
                ExpectedCellCheck(
                    name="reporting_date",
                    locations=[
                        CellLocation(
                            sheet_selector=SheetSelector(
                                mode="exact",
                                value="Data",
                            ),
                            cell_reference="A1",
                        )
                    ],
                    expected_value="2025-01-02",
                    comparison=CellComparison(
                        mode="date",
                        options={
                            "accepted_formats": ["%Y-%m-%d"],
                        },
                    ),
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True


def test_expected_cell_rejects_unknown_comparison_mode():
    with pytest.raises(ValueError):
        CellComparison(mode="semantic")


def test_validate_excel_reports_expected_cell_excluded_by_enabled_rules(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")
    config = make_config().model_copy(
        update={
            "expected_cells": [
                ExpectedCellCheck(
                    name="status",
                    locations=[
                        CellLocation(
                            sheet_selector=SheetSelector(
                                mode="exact",
                                value="Data",
                            ),
                            cell_reference="A1",
                        )
                    ],
                    expected_value="ready",
                )
            ],
            "enabled_rules": ["formula_error"],
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert len(result.not_run) == 1
    assert result.not_run[0].rule_code == "expected_cell"
    assert result.not_run[0].reason == "RULE_DISABLED"


def test_validate_excel_reports_column_rules_excluded_by_enabled_rules(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Unit"])
    workbook.active.append(["Invalid"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "enabled_rules": ["formula_error"],
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(mode="exact", value="Data"),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit"],
                    ),
                    column_validations=[
                        ColumnValidation(
                            column="unit",
                            required=True,
                            unique=True,
                            allowed_values=["£", "%"],
                            null_policy="allow",
                        )
                    ],
                )
            ],
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert {
        event.rule_code for event in result.not_run
    } == {
        "column_header",
        "missing_value",
        "duplicate_value",
        "allowed_values",
    }
    assert all(event.reason == "RULE_DISABLED" for event in result.not_run)


def test_validate_excel_reports_table_rules_excluded_by_enabled_rules(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Wrong"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "enabled_rules": ["formula_error"],
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(mode="exact", value="Data"),
                    header_row=1,
                    data_boundary=DataBoundary(
                        mode="last_non_empty_row",
                        columns=["reference"],
                    ),
                    column_definitions=[
                        ColumnDefinition(
                            name="reference",
                            aliases=["Reference"],
                        ),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["reference"],
                    ),
                    data_presence="require_one_usable_row",
                    empty_row_rules=[
                        EmptyRowRule(rows=[2], blank_policy="none_only")
                    ],
                )
            ],
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert {
        event.rule_code for event in result.not_run
    } == {"column_header", "non_empty_data", "empty_row_pattern"}
    assert all(event.reason == "RULE_DISABLED" for event in result.not_run)


def test_validate_excel_standalone_reports_missing_required_sheet(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Other")

    result = validate_excel(
        candidate_path=candidate_path,
        config=make_config(required_sheet="Data"),
    )

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["required_sheet"]
    assert result.errors[0].status == "FAILED"


def test_validate_excel_standalone_accepts_required_sheet(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")

    result = validate_excel(
        candidate_path=candidate_path,
        config=make_config(required_sheet="Data"),
    )

    assert result.is_valid is True
    assert result.errors == []


def test_validate_excel_reports_extra_sheets_as_errors(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook_with_sheets(candidate_path, ["Data", "Unexpected"])

    result = validate_excel(
        candidate_path=candidate_path,
        config=make_config(extra_sheet_action="error"),
    )

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["extra_sheet"]
    assert result.errors[0].sheet_name == "Unexpected"


def test_validate_excel_reports_extra_sheets_as_warnings(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook_with_sheets(candidate_path, ["Data", "Unexpected"])

    result = validate_excel(
        candidate_path=candidate_path,
        config=make_config(extra_sheet_action="warning"),
    )

    assert result.is_valid is True
    assert [event.rule_code for event in result.warnings] == ["extra_sheet"]


def test_validate_excel_ignores_configured_extra_sheets(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook_with_sheets(candidate_path, ["Data", "Metadata"])

    result = validate_excel(
        candidate_path=candidate_path,
        config=make_config(
            extra_sheet_action="error",
            ignored_selectors=[SheetSelector(mode="exact", value="Metadata")],
        ),
    )

    assert result.is_valid is True
    assert result.errors == []


def test_validate_excel_reports_disabled_workbook_check(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")
    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_error",
                enabled=False,
                scope="all_sheets",
            )
        ]
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert result.errors == []
    assert len(result.not_run) == 1
    assert result.not_run[0].rule_code == "formula_error"
    assert result.not_run[0].reason == "RULE_DISABLED"


def test_validate_excel_reports_rule_excluded_by_enabled_rules(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")
    config = make_config(
        workbook_checks=[
                WorkbookCheck(
                    rule_code="formula_error",
                    enabled=True,
                    scope="all_sheets",
                    options={"error_tokens": ["#DIV/0!"]},
                )
        ]
    ).model_copy(update={"enabled_rules": []})

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert len(result.not_run) == 1
    assert result.not_run[0].rule_code == "formula_error"
    assert result.not_run[0].reason == "RULE_DISABLED"


def test_validate_excel_reports_configured_formula_errors(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = "#DIV/0!"
    workbook.save(candidate_path)

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_error",
                enabled=True,
                scope="all_sheets",
                options={"error_tokens": ["#DIV/0!"]},
            )
        ]
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert len(result.errors) == 1
    assert result.errors[0].rule_code == "formula_error"
    assert result.errors[0].sheet_name == "Data"
    assert result.errors[0].cell_reference == "A1"


def test_formula_error_is_not_run_when_formula_cache_is_missing(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = "=1+1"
    workbook.save(candidate_path)

    result = validate_excel(
        candidate_path=candidate_path,
        config=make_config(
            workbook_checks=[
                WorkbookCheck(
                    rule_code="formula_error",
                    enabled=True,
                    scope="all_sheets",
                    options={"error_tokens": ["#DIV/0!"]},
                )
            ]
        ),
    )

    assert result.is_valid is True
    assert len(result.not_run) == 1
    assert result.not_run[0].rule_code == "formula_error"
    assert result.not_run[0].reason == "FORMULA_RESULTS_UNAVAILABLE"
    assert result.not_run[0].cell_reference == "A1"


def test_formula_error_fails_when_formula_cache_is_missing_and_policy_is_error(
    tmp_path,
):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = "=1+1"
    workbook.save(candidate_path)

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_error",
                enabled=True,
                scope="all_sheets",
                options={"error_tokens": ["#DIV/0!"]},
            )
        ]
    ).model_copy(
        update={
            "runtime": make_config().runtime.model_copy(
                update={"missing_formula_cache_action": "error"}
            )
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert len(result.errors) == 1
    assert result.errors[0].rule_code == "formula_error"
    assert result.errors[0].reason == "FORMULA_RESULTS_UNAVAILABLE"
    assert result.errors[0].cell_reference == "A1"


def test_validate_excel_reports_missing_reference_sheet(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    reference_path = tmp_path / "reference.xlsx"
    write_workbook_with_sheets(candidate_path, ["Data"])
    write_workbook_with_sheets(reference_path, ["Data", "Required_From_Template"])

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="missing_reference_sheet",
                enabled=True,
                scope="all_sheets",
            )
        ]
    )

    result = validate_excel(
        candidate_path=candidate_path,
        reference_path=reference_path,
        config=config,
    )

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == [
        "missing_reference_sheet"
    ]
    assert result.errors[0].sheet_name == "Required_From_Template"


def test_validate_excel_reports_formula_differences(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    reference_path = tmp_path / "reference.xlsx"

    candidate = openpyxl.Workbook()
    candidate.active.title = "Data"
    candidate.active["B1"] = "=SUM(A1)"
    candidate.save(candidate_path)

    reference = openpyxl.Workbook()
    reference.active.title = "Data"
    reference.active["B1"] = "=SUM(A2)"
    reference.save(reference_path)

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_difference",
                enabled=True,
                scope="overlapping_sheets",
            )
        ]
    )

    result = validate_excel(
        candidate_path=candidate_path,
        reference_path=reference_path,
        config=config,
    )

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == [
        "formula_difference"
    ]
    assert result.errors[0].sheet_name == "Data"
    assert result.errors[0].cell_reference == "B1"
    assert result.errors[0].actual_value == "=SUM(A1)"
    assert result.errors[0].expected_value == "=SUM(A2)"


def test_reference_formula_check_is_not_run_without_reference_workbook(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")
    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_difference",
                enabled=True,
                scope="overlapping_sheets",
            )
        ]
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert len(result.not_run) == 1
    assert result.not_run[0].reason == "REFERENCE_WORKBOOK_REQUIRED"


def test_validate_excel_accepts_identical_array_formulas(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    reference_path = tmp_path / "reference.xlsx"

    for path in (candidate_path, reference_path):
        workbook = openpyxl.Workbook()
        workbook.active.title = "Data"
        workbook.active["A3"] = ArrayFormula(ref="A3", text="=SUM(A1:A2)")
        workbook.save(path)

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_difference",
                enabled=True,
                scope="overlapping_sheets",
            )
        ]
    )

    result = validate_excel(
        candidate_path=candidate_path,
        reference_path=reference_path,
        config=config,
    )

    assert result.is_valid is True
    assert result.errors == []


def test_validate_excel_ignores_formula_whitespace_differences(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    reference_path = tmp_path / "reference.xlsx"

    for path, formula in (
        (candidate_path, "= 1+1"),
        (reference_path, "=1+1"),
    ):
        workbook = openpyxl.Workbook()
        workbook.active.title = "Data"
        workbook.active["A1"] = formula
        workbook.save(path)

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_difference",
                enabled=True,
                scope="overlapping_sheets",
            )
        ]
    )

    result = validate_excel(
        candidate_path=candidate_path,
        reference_path=reference_path,
        config=config,
    )

    assert result.is_valid is True
    assert result.errors == []


def test_validate_excel_supports_exact_formula_comparison(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    reference_path = tmp_path / "reference.xlsx"

    for path, formula in (
        (candidate_path, "= 1+1"),
        (reference_path, "=1+1"),
    ):
        workbook = openpyxl.Workbook()
        workbook.active.title = "Data"
        workbook.active["A1"] = formula
        workbook.save(path)

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_difference",
                enabled=True,
                scope="overlapping_sheets",
                options={"whitespace_policy": "exact"},
            )
        ]
    )

    result = validate_excel(
        candidate_path=candidate_path,
        reference_path=reference_path,
        config=config,
    )

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == [
        "formula_difference"
    ]


def test_validate_excel_reports_expected_cell_difference(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")
    workbook = openpyxl.load_workbook(candidate_path)
    workbook["Data"]["A1"] = "actual"
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "expected_cells": [
                ExpectedCellCheck(
                    name="status",
                    locations=[
                        CellLocation(
                            sheet_selector=SheetSelector(
                                mode="exact",
                                value="Data",
                            ),
                            cell_reference="A1",
                        )
                    ],
                    expected_value="expected",
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["expected_cell"]
    assert result.errors[0].cell_reference == "A1"
    assert result.errors[0].actual_value == "actual"
    assert result.errors[0].expected_value == "expected"


def test_validate_excel_reports_missing_table_header(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Unit", "Other"])
    workbook.active["Z11"]
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    required=True,
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                        ColumnDefinition(
                            name="reference",
                            aliases=["Reference"],
                        ),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit", "reference"],
                        match_mode="contains_in_order",
                        extra_column_action="allowed",
                        missing_column_action="error",
                    ),
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["column_header"]
    assert result.errors[0].sheet_name == "Data"
    assert result.errors[0].actual_value == ["Unit", "Other"]


def test_validate_excel_reports_missing_required_table(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Outputs",
                    ),
                    required=True,
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="reference"),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["reference"],
                    ),
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["table_resolution"]
    assert result.errors[0].description == (
        "Required table 'outputs' did not resolve"
    )


def test_validate_excel_does_not_validate_tables_on_ignored_sheets(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook_with_sheets(candidate_path, ["Data", "Metadata"])

    config = make_config(
        ignored_selectors=[
            SheetSelector(mode="exact", value="Metadata"),
        ],
    ).model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Metadata",
                    ),
                    required=True,
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="reference"),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["reference"],
                    ),
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["table_resolution"]


def test_sheet_structure_ignores_blank_cells_outside_content(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    reference_path = tmp_path / "reference.xlsx"

    candidate = openpyxl.Workbook()
    candidate.active.title = "Data"
    candidate.active["A1"] = "value"
    candidate.active["Z11"].fill = PatternFill(
        fill_type="solid",
        fgColor="FFFF00",
    )
    candidate.save(candidate_path)

    reference = openpyxl.Workbook()
    reference.active.title = "Data"
    reference.active["A1"] = "value"
    reference.save(reference_path)

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="sheet_structure",
                enabled=True,
                scope="overlapping_sheets",
            )
        ]
    )

    result = validate_excel(
        candidate_path=candidate_path,
        reference_path=reference_path,
        config=config,
    )

    assert result.is_valid is True


def test_expected_cell_reports_excel_error_value_mismatch(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = "#DIV/0!"
    workbook.save(candidate_path)

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_error",
                enabled=True,
                scope="all_sheets",
                options={"error_tokens": ["#DIV/0!"]},
            )
        ],
    ).model_copy(
        update={
            "expected_cells": [
                ExpectedCellCheck(
                    name="value",
                    locations=[
                        CellLocation(
                            sheet_selector=SheetSelector(
                                mode="exact",
                                value="Data",
                            ),
                            cell_reference="A1",
                        )
                    ],
                    expected_value="expected",
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert [event.rule_code for event in result.errors] == [
        "expected_cell",
        "formula_error",
    ]
    assert result.errors[0].actual_value == "#DIV/0!"
    assert result.errors[0].expected_value == "expected"
    assert result.not_run == []


def test_validate_excel_reports_required_table_without_data(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Unit", "Reference"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    required=True,
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                        ColumnDefinition(
                            name="reference",
                            aliases=["Reference"],
                        ),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit", "reference"],
                    ),
                    data_presence="require_one_usable_row",
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["non_empty_data"]
    assert result.errors[0].sheet_name == "Data"


def test_validate_excel_reports_disallowed_column_values(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Unit"])
    workbook.active.append(["£"])
    workbook.active.append(["GBP"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit"],
                        match_mode="contains_any_order",
                    ),
                    column_validations=[
                        ColumnValidation(
                            column="unit",
                            allowed_values=["£", "%"],
                            null_policy="allow",
                        )
                    ],
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["allowed_values"]
    assert result.errors[0].cell_reference == "A3"
    assert result.errors[0].actual_value == "GBP"


def test_column_values_use_case_and_whitespace_normalization(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Unit", "Reference"])
    workbook.active.append([" ml ", "ABC"])
    workbook.active.append(["Ml", " abc "])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    data_boundary=DataBoundary(
                        mode="last_non_empty_row",
                        columns=["reference"],
                    ),
                    column_definitions=[
                        ColumnDefinition(name="unit"),
                        ColumnDefinition(name="reference"),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit", "reference"],
                    ),
                    column_validations=[
                        ColumnValidation(
                            column="unit",
                            allowed_values=["Ml"],
                        ),
                        ColumnValidation(
                            column="reference",
                            unique=True,
                        ),
                    ],
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert [event.rule_code for event in result.errors] == ["duplicate_value"]
    assert result.errors[0].cell_reference == "B3"


def test_column_values_use_accent_normalization_when_configured(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Status"])
    workbook.active.append(["cafe"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "comparison": ComparisonConfig(
                case_sensitive=False,
                accent_sensitive=False,
                trim_whitespace=True,
                collapse_internal_whitespace=False,
                empty_string_is_null=True,
                null_tokens=[],
                collation_name="mssql_case_insensitive_ai",
            ),
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="status"),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["status"],
                    ),
                    column_validations=[
                        ColumnValidation(
                            column="status",
                            allowed_values=["café"],
                        )
                    ],
                )
            ],
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True


def test_null_tokens_use_configured_case_and_whitespace_policy(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Reference"])
    workbook.active.append(["N/A"])
    workbook.active.append([" n/a "])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "comparison": ComparisonConfig(
                case_sensitive=False,
                accent_sensitive=True,
                trim_whitespace=True,
                collapse_internal_whitespace=False,
                empty_string_is_null=True,
                null_tokens=["N/A"],
                collation_name="mssql_case_insensitive",
            ),
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="reference"),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["reference"],
                    ),
                    column_validations=[
                        ColumnValidation(
                            column="reference",
                            unique=True,
                            null_policy="allow",
                        )
                    ],
                )
            ],
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True


def test_column_definition_can_override_comparison_policy(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Status"])
    workbook.active.append(["active"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "comparison": ComparisonConfig(
                case_sensitive=True,
                accent_sensitive=True,
                trim_whitespace=True,
                collapse_internal_whitespace=False,
                empty_string_is_null=True,
                null_tokens=[],
                collation_name="case_sensitive",
            ),
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(
                            name="status",
                            comparison=ComparisonConfig(
                                case_sensitive=False,
                                accent_sensitive=True,
                                trim_whitespace=True,
                                collapse_internal_whitespace=False,
                                empty_string_is_null=True,
                                null_tokens=[],
                                collation_name="case_insensitive",
                            ),
                        ),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["status"],
                    ),
                    column_validations=[
                        ColumnValidation(
                            column="status",
                            allowed_values=["Active"],
                        )
                    ],
                )
            ],
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True


def test_column_rules_are_not_run_when_headers_cannot_resolve(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Wrong"])
    workbook.active.append(["GBP"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit"],
                        match_mode="contains_any_order",
                    ),
                    column_validations=[
                        ColumnValidation(
                            column="unit",
                            allowed_values=["£", "%"],
                            null_policy="allow",
                        )
                    ],
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert [event.rule_code for event in result.errors] == ["column_header"]
    assert len(result.not_run) == 1
    assert result.not_run[0].rule_code == "allowed_values"
    assert result.not_run[0].reason == "TABLE_HEADER_RESOLUTION_FAILED"


def test_validate_excel_reports_required_and_duplicate_column_values(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Unit"])
    workbook.active.append([None])
    workbook.active.append(["£"])
    workbook.active.append(["£"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit"],
                    ),
                    column_validations=[
                        ColumnValidation(
                            column="unit",
                            required=True,
                            unique=True,
                            null_policy="allow",
                        )
                    ],
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == [
        "missing_value",
        "duplicate_value",
    ]
    assert result.errors[0].cell_reference == "A2"
    assert result.errors[1].cell_reference == "A4"


def test_validate_excel_reports_sheet_structure_differences(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    reference_path = tmp_path / "reference.xlsx"

    candidate = openpyxl.Workbook()
    candidate.active.title = "Data"
    candidate.active["A1"] = "value"
    candidate.save(candidate_path)

    reference = openpyxl.Workbook()
    reference.active.title = "Data"
    reference.active["B1"] = "value"
    reference.save(reference_path)

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="sheet_structure",
                enabled=True,
                scope="overlapping_sheets",
            )
        ]
    )

    result = validate_excel(
        candidate_path=candidate_path,
        reference_path=reference_path,
        config=config,
    )

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["sheet_structure"]
    assert result.errors[0].sheet_name == "Data"
    assert result.errors[0].actual_value == {
        "max_row": 1,
        "max_column": 1,
    }
    assert result.errors[0].expected_value == {
        "max_row": 1,
        "max_column": 2,
    }


def test_reference_structure_check_is_not_run_without_reference_workbook(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")
    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="sheet_structure",
                enabled=True,
                scope="overlapping_sheets",
            )
        ]
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert len(result.not_run) == 1
    assert result.not_run[0].reason == "REFERENCE_WORKBOOK_REQUIRED"


def test_reference_rule_is_not_run_without_reference_workbook(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data")
    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="missing_reference_sheet",
                enabled=True,
                scope="all_sheets",
            )
        ]
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert len(result.not_run) == 1
    assert result.not_run[0].reason == "REFERENCE_WORKBOOK_REQUIRED"


def test_validate_excel_reports_non_blank_configured_empty_row(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Unit", "Reference"])
    workbook.active.append(["£", "A"])
    workbook.active.append([None, "Unexpected"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                        ColumnDefinition(
                            name="reference",
                            aliases=["Reference"],
                        ),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit", "reference"],
                    ),
                    empty_row_rules=[
                        EmptyRowRule(
                            rows=[3],
                            excluded_columns=["A"],
                            blank_policy="blank_strings",
                        )
                    ],
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == [
        "empty_row_pattern"
    ]
    assert result.errors[0].cell_reference == "B3"


def test_validate_excel_enforces_exact_header_order(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Reference", "Unit"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(mode="exact", value="Data"),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                        ColumnDefinition(
                            name="reference",
                            aliases=["Reference"],
                        ),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit", "reference"],
                        match_mode="exact_order",
                    ),
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["column_header"]


def test_validate_excel_accepts_any_header_order(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Reference", "Unit"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(mode="exact", value="Data"),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                        ColumnDefinition(
                            name="reference",
                            aliases=["Reference"],
                        ),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit", "reference"],
                        match_mode="contains_any_order",
                    ),
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True


def test_validate_excel_reports_extra_table_headers(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Unit", "Extra"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(mode="exact", value="Data"),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit"],
                        extra_column_action="error",
                    ),
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["column_header"]
    assert result.errors[0].actual_value == "Extra"


def test_validate_excel_reports_duplicate_table_headers(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Unit", "Unit"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(mode="exact", value="Data"),
                    header_row=1,
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit"],
                        match_mode="contains_any_order",
                    ),
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["column_header"]
    assert "duplicate" in result.errors[0].description


def test_column_rules_respect_excel_table_data_boundary(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    sheet = workbook.active
    sheet.append(["Reference", "Status"])
    sheet.append(["A", "Valid"])
    sheet.append(["B", "Invalid"])
    sheet.add_table(Table(displayName="Outputs", ref="A1:B2"))
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    data_boundary=DataBoundary(
                        mode="excel_table",
                        table_name="Outputs",
                    ),
                    column_definitions=[
                        ColumnDefinition(name="status"),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["reference", "status"],
                    ),
                    column_validations=[
                        ColumnValidation(
                            column="status",
                            forbidden_values=["Invalid"],
                        )
                    ],
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True


def test_missing_excel_table_boundary_is_reported(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Reference"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    data_boundary=DataBoundary(
                        mode="excel_table",
                        table_name="MissingTable",
                    ),
                    column_definitions=[
                        ColumnDefinition(name="reference"),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["reference"],
                    ),
                    data_presence="require_one_usable_row",
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert [event.rule_code for event in result.errors] == [
        "table_resolution",
    ]
    assert result.not_run[0].rule_code == "non_empty_data"


def test_column_rules_respect_fixed_table_data_boundary(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Unit"])
    workbook.active.append(["Invalid"])
    workbook.active.append(["£"])
    workbook.active.append(["£"])
    workbook.active.append(["Invalid_After_Boundary"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    data_boundary=DataBoundary(
                        mode="fixed_end_row",
                        end_row=4,
                    ),
                    column_definitions=[
                        ColumnDefinition(name="unit", aliases=["Unit"]),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["unit"],
                    ),
                    column_validations=[
                        ColumnValidation(
                            column="unit",
                            allowed_values=["£", "%"],
                            null_policy="allow",
                        )
                    ],
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert [event.rule_code for event in result.errors] == ["allowed_values"]
    assert result.errors[0].cell_reference == "A2"


def test_data_presence_is_not_run_when_boundary_column_cannot_resolve(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Wrong"])
    workbook.active.append(["value"])
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "tables": [
                TableConfig(
                    name="outputs",
                    sheet_selector=SheetSelector(
                        mode="exact",
                        value="Data",
                    ),
                    header_row=1,
                    data_boundary=DataBoundary(
                        mode="last_non_empty_row",
                        columns=["reference"],
                    ),
                    column_definitions=[
                        ColumnDefinition(name="reference", aliases=["Reference"]),
                    ],
                    header_policy=HeaderPolicy(
                        required_columns=["reference"],
                    ),
                    data_presence="require_one_usable_row",
                )
            ]
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert [event.rule_code for event in result.errors] == ["column_header"]
    assert len(result.not_run) == 1
    assert result.not_run[0].rule_code == "non_empty_data"
    assert result.not_run[0].reason == "DATA_BOUNDARY_RESOLUTION_FAILED"


def test_validate_excel_stops_when_cell_scan_limit_is_exceeded(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = "first"
    workbook.active["A2"] = "second"
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "runtime": RuntimePolicy(
                max_cells_scanned=1,
                read_only=False,
                macro_policy="reject",
                missing_formula_cache_action="not_run",
            )
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert result.complete is False
    assert [event.rule_code for event in result.errors] == [
        "WORKBOOK_LIMIT_ERROR"
    ]
    assert result.errors[0].actual_value == 2
    assert result.errors[0].expected_value == 1


def test_scan_limit_counts_stored_cells_not_inflated_dimensions(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1048576"] = "tail"
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "runtime": RuntimePolicy(
                max_cells_scanned=1,
                read_only=False,
                macro_policy="reject",
                missing_formula_cache_action="not_run",
            )
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert result.complete is True


def test_read_only_workbook_validation_uses_public_api(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = "#DIV/0!"
    workbook.save(candidate_path)

    config = make_config(
        workbook_checks=[
            WorkbookCheck(
                rule_code="formula_error",
                enabled=True,
                scope="all_sheets",
                options={"error_tokens": ["#DIV/0!"]},
            )
        ]
    ).model_copy(
        update={
            "runtime": RuntimePolicy(
                max_cells_scanned=1000,
                read_only=True,
                macro_policy="reject",
                missing_formula_cache_action="not_run",
            )
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert result.errors[0].cell_reference == "A1"


def test_feature_policy_reports_merged_cells(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.merge_cells("A1:B1")
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "runtime": RuntimePolicy(
                max_cells_scanned=1000,
                read_only=False,
                macro_policy="reject",
                missing_formula_cache_action="not_run",
                feature_policy=WorkbookFeaturePolicy(
                    merged_cells="error",
                ),
            )
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["feature_policy"]
    assert "merged_cells" in result.errors[0].description


def test_feature_policy_warning_does_not_invalidate_result(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.merge_cells("A1:B1")
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "runtime": RuntimePolicy(
                max_cells_scanned=1000,
                read_only=False,
                macro_policy="reject",
                missing_formula_cache_action="not_run",
                feature_policy=WorkbookFeaturePolicy(
                    merged_cells="warning",
                ),
            )
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert [event.rule_code for event in result.warnings] == ["feature_policy"]
    assert result.diagnostics == []


def test_feature_policy_ignore_records_a_diagnostic(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active.merge_cells("A1:B1")
    workbook.save(candidate_path)

    config = make_config().model_copy(
        update={
            "runtime": RuntimePolicy(
                max_cells_scanned=1000,
                read_only=False,
                macro_policy="reject",
                missing_formula_cache_action="not_run",
                feature_policy=WorkbookFeaturePolicy(
                    merged_cells="ignore",
                ),
            )
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert result.warnings == []
    assert [event.code for event in result.diagnostics] == [
        "FEATURE_DETECTED",
    ]


def test_validate_excel_supports_regex_sheet_selectors(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook(candidate_path, "Data_2025")

    config = make_config(required_sheet="^Data_[0-9]{4}$").model_copy(
        update={
            "sheet_policy": SheetPolicy(
                required_selectors=[
                    SheetSelector(
                        mode="regex",
                        value="^Data_[0-9]{4}$",
                    )
                ],
                ignored_selectors=[],
                extra_sheet_action="ignore",
                selector_match_action="all",
            )
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is True
    assert result.errors == []


def test_validate_excel_reports_multiple_selector_matches(tmp_path):
    candidate_path = tmp_path / "candidate.xlsx"
    write_workbook_with_sheets(candidate_path, ["Data_2024", "Data_2025"])

    config = make_config().model_copy(
        update={
            "sheet_policy": SheetPolicy(
                required_selectors=[
                    SheetSelector(
                        mode="regex",
                        value="^Data_[0-9]{4}$",
                    )
                ],
                ignored_selectors=[],
                extra_sheet_action="ignore",
                selector_match_action="error_on_multiple",
            )
        }
    )

    result = validate_excel(candidate_path=candidate_path, config=config)

    assert result.is_valid is False
    assert [event.rule_code for event in result.errors] == ["sheet_selector"]
    assert result.errors[0].actual_value == 2
