from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass, field
import re
import unicodedata
from pathlib import Path
from zipfile import ZipFile

import openpyxl
import pandas as pd
from openpyxl.utils.cell import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)
from openpyxl.worksheet._reader import WorkSheetParser

from .config import validate_config
from .models import (
    ConfigurationError,
    ColumnDefinition,
    DiagnosticEvent,
    RuleContext,
    ValidationEvent,
    ValidationResult,
    WorkbookReadError,
    WorkbookValidationConfig,
)
from .registry import get_registered_rule, validate_custom_result


def _ordered_workbook_checks(checks):
    by_rule = {check.rule_code: check for check in checks}
    ordered = []
    visited = set()

    def visit(rule_code):
        if rule_code in visited:
            return
        visited.add(rule_code)
        check = by_rule[rule_code]
        for dependency in check.depends_on:
            visit(dependency)
        ordered.append(check)

    for check in checks:
        visit(check.rule_code)
    return ordered


def validate_excel(
    candidate_path,
    config: WorkbookValidationConfig,
    reference_path=None,
    candidate_version=None,
    reference_version=None,
    run_id=None,
    profile=None,
):
    resolved_config = validate_config(config)
    candidate_path = Path(candidate_path)
    reference_path = Path(reference_path) if reference_path is not None else None
    result_metadata = {
        "run_id": run_id,
        "config_version": resolved_config.config_version,
        "profile_name": resolved_config.profile_name,
        "candidate_filename": candidate_path.name,
        "candidate_version": candidate_version,
        "reference_filename": reference_path.name if reference_path else None,
        "reference_version": reference_version,
        "comparison": resolved_config.comparison,
    }
    formula_workbook = None
    value_workbook = None
    reference_formula_workbook = None
    reference_workbook = None
    try:
        formula_workbook = openpyxl.load_workbook(
            candidate_path,
            read_only=resolved_config.runtime.read_only,
            data_only=False,
            keep_vba=candidate_path.suffix.lower() == ".xlsm",
        )
        value_workbook = openpyxl.load_workbook(
            candidate_path,
            read_only=resolved_config.runtime.read_only,
            data_only=True,
            keep_vba=candidate_path.suffix.lower() == ".xlsm",
        )
        if reference_path is not None:
            reference_formula_workbook = openpyxl.load_workbook(
                reference_path,
                read_only=resolved_config.runtime.read_only,
                data_only=False,
                keep_vba=reference_path.suffix.lower() == ".xlsm",
            )
            reference_workbook = openpyxl.load_workbook(
                reference_path,
                read_only=resolved_config.runtime.read_only,
                data_only=True,
                keep_vba=Path(reference_path).suffix.lower() == ".xlsm",
            )
    except Exception as error:
        if formula_workbook is not None:
            formula_workbook.close()
        if value_workbook is not None:
            value_workbook.close()
        if reference_workbook is not None:
            reference_workbook.close()
        if reference_formula_workbook is not None:
            reference_formula_workbook.close()
        raise WorkbookReadError(str(error)) from error

    try:
        errors = []
        warnings = []
        not_run = []
        diagnostics = []
        processed_counts = {}
        processed_severities = dict(resolved_config.rule_severity)
        for check in resolved_config.workbook_checks:
            processed_severities.setdefault(
                check.rule_code,
                check.severity,
            )
        processed_severities.setdefault(
            "extra_sheet",
            resolved_config.sheet_policy.extra_sheet_action,
        )

        def record_processed(rule_code=None, count=1, code=None):
            key = f"code:{code}" if code is not None else rule_code
            if key is not None:
                processed_counts[key] = processed_counts.get(key, 0) + count

        def finalize_processed_counts():
            for event in [*errors, *warnings, *not_run]:
                processed_counts.setdefault(event.rule_code, 0)
            for diagnostic in diagnostics:
                processed_counts.setdefault(f"code:{diagnostic.code}", 0)
        for (
            feature_name,
            sheet_name,
            cell_reference,
            detected_value,
            detection_error,
        ) in _detected_features(
            formula_workbook,
            candidate_path,
        ):
            if detection_error is not None:
                action = resolved_config.runtime.feature_policy.unavailable_action
                event_code = f"{feature_name}_detection_unavailable"
                if action == "ignore":
                    processed_severities[f"code:{event_code.upper()}"] = action
                    record_processed(code=event_code.upper())
                else:
                    processed_severities[event_code] = action
                    record_processed(event_code)
                description = (
                    f"Feature detection unavailable for {feature_name}: "
                    f"{detection_error}"
                )
                event = ValidationEvent(
                    rule_code=event_code,
                    reason=event_code.upper(),
                    sheet_name=sheet_name,
                    cell_reference=cell_reference,
                    actual_value=detected_value,
                    severity=action,
                    description=description,
                )
                if action == "error":
                    errors.append(event)
                elif action == "warning":
                    warnings.append(event)
                else:
                    diagnostics.append(
                        DiagnosticEvent(
                            run_id=run_id,
                            code=event_code.upper(),
                            description=description,
                        )
                    )
                continue
            action = getattr(
                resolved_config.runtime.feature_policy,
                feature_name,
            )
            event_code = feature_name
            if action == "ignore":
                processed_severities[f"code:{event_code.upper()}"] = action
                record_processed(code=event_code.upper())
            else:
                processed_severities[event_code] = action
                record_processed(event_code)
            description = f"Detected workbook feature: {feature_name}"
            if cell_reference is not None:
                description += f" ({cell_reference})"
            if detected_value is not None:
                description += f": {detected_value}"
            event = ValidationEvent(
                rule_code=event_code,
                reason=f"{event_code.upper()}_DETECTED",
                sheet_name=sheet_name,
                cell_reference=cell_reference,
                actual_value=detected_value,
                severity=action,
                description=description,
            )
            if action == "error":
                errors.append(event)
            elif action == "warning":
                warnings.append(event)
            else:
                diagnostics.append(
                    DiagnosticEvent(
                        run_id=run_id,
                        code=f"{event_code.upper()}_DETECTED",
                        description=description,
                    )
                )
        sheet_names = set(formula_workbook.sheetnames)
        selected_sheet_names = set()
        for selector in resolved_config.sheet_policy.required_selectors:
            record_processed("required_sheet")
            record_processed("sheet_selector")
            matches = _select_sheet_matches(
                _match_sheets(sheet_names, selector),
                resolved_config.sheet_policy.selector_match_action,
                selector,
                errors,
            )
            selected_sheet_names.update(matches)
            if not matches:
                errors.append(
                    ValidationEvent(
                        rule_code="required_sheet",
                        sheet_name=selector.value,
                        description=(
                            f"Required sheet selector did not match: "
                            f"{selector.value}"
                        ),
                    )
                )
        for selector in resolved_config.sheet_policy.ignored_selectors:
            record_processed("sheet_selector")
            selected_sheet_names.update(
                _select_sheet_matches(
                    _match_sheets(sheet_names, selector),
                    resolved_config.sheet_policy.selector_match_action,
                    selector,
                    errors,
                )
            )
        ignored_sheet_names = set()
        for selector in resolved_config.sheet_policy.ignored_selectors:
            ignored_sheet_names.update(
                _match_sheets(sheet_names, selector)
            )
        if (
            not resolved_config.sheet_policy.required_selectors
            and reference_workbook is not None
        ):
            selected_sheet_names.update(reference_workbook.sheetnames)
        extra_sheets = sheet_names - selected_sheet_names
        action = resolved_config.sheet_policy.extra_sheet_action
        for sheet_name in formula_workbook.sheetnames:
            record_processed("extra_sheet")
            if sheet_name not in extra_sheets or action == "ignore":
                continue
            event = ValidationEvent(
                rule_code="extra_sheet",
                severity=action,
                sheet_name=sheet_name,
                description=f"Extra sheet found: {sheet_name}",
            )
            if action == "warning":
                warnings.append(event)
            elif action == "error":
                errors.append(event)
        value_resolution = _WorkbookResolution(value_workbook)
        formula_resolution = _WorkbookResolution(formula_workbook)
        reference_value_resolution = (
            _WorkbookResolution(reference_workbook)
            if reference_workbook is not None
            else None
        )
        reference_formula_resolution = (
            _WorkbookResolution(reference_formula_workbook)
            if reference_formula_workbook is not None
            else None
        )
        observed_cells = _cells_until_limit(
            value_workbook,
            resolved_config.runtime.max_cells_scanned,
            value_resolution,
        )
        if observed_cells is not None:
            record_processed("WORKBOOK_LIMIT_ERROR", observed_cells)
            errors.append(
                ValidationEvent(
                    rule_code="WORKBOOK_LIMIT_ERROR",
                    actual_value=observed_cells,
                    expected_value=resolved_config.runtime.max_cells_scanned,
                    description=(
                        "Workbook cell scan limit exceeded; validation stopped"
                    ),
                )
            )
            finalize_processed_counts()
            return ValidationResult(
                **result_metadata,
                complete=False,
                processed_counts=processed_counts,
                processed_severities=processed_severities,
                errors=errors,
                warnings=warnings,
                not_run=not_run,
                diagnostics=diagnostics,
            )
        for expected_cell in resolved_config.expected_cells:
            record_processed("expected_cell")
            if (
                not expected_cell.enabled
                or (
                    resolved_config.enabled_rules is not None
                    and "expected_cell" not in resolved_config.enabled_rules
                )
            ):
                not_run.append(
                    ValidationEvent(
                        rule_code="expected_cell",
                        status="NOT_RUN",
                        reason="RULE_DISABLED",
                        severity=expected_cell.severity
                        or resolved_config.rule_severity.get("expected_cell"),
                        description="Rule is disabled by configuration",
                    )
                )
                continue
            resolved_locations = []
            for location in expected_cell.locations:
                matches = _match_sheets(
                    value_workbook.sheetnames,
                    location.sheet_selector,
                )
                resolved_locations.extend(
                    (sheet_name, location.cell_reference) for sheet_name in matches
                )
                if (
                    expected_cell.fallback_strategy == "ordered_first"
                    and resolved_locations
                ):
                    break
            if not resolved_locations:
                errors.append(
                    ValidationEvent(
                        rule_code="expected_cell",
                        severity=expected_cell.severity
                        or resolved_config.rule_severity.get("expected_cell"),
                        description=(
                            f"Expected cell location did not resolve: "
                            f"{expected_cell.name}"
                        ),
                    )
                )
                continue
            if (
                expected_cell.fallback_strategy == "exactly_one"
                and len(resolved_locations) != 1
            ):
                errors.append(
                    ValidationEvent(
                        rule_code="expected_cell",
                        severity=expected_cell.severity
                        or resolved_config.rule_severity.get("expected_cell"),
                        description=(
                            f"Expected cell requires exactly one location: "
                            f"{expected_cell.name}"
                        ),
                    )
                )
                continue
            for sheet_name, cell_reference in resolved_locations:
                actual_cell = value_workbook[sheet_name][cell_reference]
                actual_value = actual_cell.value
                if _values_equal(
                    actual_value,
                    expected_cell.expected_value,
                    expected_cell.comparison.mode,
                    expected_cell.comparison.options,
                ):
                    continue
                errors.append(
                    ValidationEvent(
                        rule_code="expected_cell",
                        severity=expected_cell.severity
                        or resolved_config.rule_severity.get("expected_cell"),
                        sheet_name=sheet_name,
                        cell_reference=cell_reference,
                        actual_value=actual_value,
                        expected_value=expected_cell.expected_value,
                        description=(
                            f"Expected cell value differs: {expected_cell.name}"
                        ),
                    )
                )
                if expected_cell.fallback_strategy == "ordered_first":
                    break
        for table in resolved_config.tables:
            if table.header_policy is None or table.sheet_selector is None:
                continue
            record_processed("table_resolution")
            record_processed("column_header")
            if table.data_presence == "require_one_usable_row":
                record_processed("non_empty_data")
            for empty_row_rule in table.empty_row_rules:
                record_processed("empty_row_pattern", len(empty_row_rule.rows))
            sheet_matches = _match_sheets(
                value_workbook.sheetnames,
                table.sheet_selector,
            )
            sheet_matches = [
                sheet_name
                for sheet_name in sheet_matches
                if sheet_name not in ignored_sheet_names
            ]
            if not sheet_matches:
                if table.required:
                    event = ValidationEvent(
                        rule_code="table_resolution",
                        status=(
                            "FAILED"
                            if (
                                resolved_config.enabled_rules is None
                                or "table_resolution"
                                in resolved_config.enabled_rules
                            )
                            else "NOT_RUN"
                        ),
                        reason=(
                            None
                            if (
                                resolved_config.enabled_rules is None
                                or "table_resolution"
                                in resolved_config.enabled_rules
                            )
                            else "RULE_DISABLED"
                        ),
                        sheet_name=table.sheet_selector.value,
                        severity=resolved_config.rule_severity.get(
                            "table_resolution"
                        ),
                        description=(
                            f"Required table '{table.name}' did not resolve"
                        ),
                    )
                    if event.status == "NOT_RUN":
                        not_run.append(event)
                    else:
                        errors.append(event)
                continue
            if table.header_row is None:
                continue
            definitions = {definition.name: definition for definition in table.column_definitions}
            for sheet_name in sheet_matches:
                enabled_rules = resolved_config.enabled_rules
                header_rule_enabled = (
                    enabled_rules is None or "column_header" in enabled_rules
                )
                data_presence_rule_enabled = (
                    table.data_presence != "require_one_usable_row"
                    or enabled_rules is None
                    or "non_empty_data" in enabled_rules
                )
                empty_row_rule_enabled = (
                    not table.empty_row_rules
                    or enabled_rules is None
                    or "empty_row_pattern" in enabled_rules
                )
                if not header_rule_enabled:
                    not_run.append(
                        ValidationEvent(
                            rule_code="column_header",
                            status="NOT_RUN",
                            reason="RULE_DISABLED",
                            description=(
                                f"Header checks for table '{table.name}' "
                                "are disabled by configuration"
                            ),
                        )
                    )
                sheet = value_workbook[sheet_name]
                headers = value_resolution.headers(
                    sheet,
                    table.header_row,
                    table,
                )
                while headers and headers[-1] is None:
                    headers.pop()
                physical_header_issues = []
                seen_headers = set()
                for header in headers:
                    if header is None or (
                        isinstance(header, str) and not header.strip()
                    ):
                        physical_header_issues.append("blank header")
                        continue
                    normalised_header = _normalise_header(header)
                    if normalised_header in seen_headers:
                        physical_header_issues.append(
                            f"duplicate header: {header}"
                        )
                    seen_headers.add(normalised_header)
                if physical_header_issues and header_rule_enabled:
                    errors.append(
                        ValidationEvent(
                            rule_code="column_header",
                            sheet_name=sheet_name,
                            actual_value=headers,
                            description=(
                                f"Table '{table.name}' has invalid physical "
                                f"headers: {', '.join(physical_header_issues)}"
                            ),
                        )
                    )
                normalised_headers = {
                    _normalise_header(header): index
                    for index, header in enumerate(headers)
                    if header is not None
                }
                logical_columns = {}
                for definition in table.column_definitions:
                    names = [definition.name, *definition.aliases]
                    for name in names:
                        position = normalised_headers.get(_normalise_header(name))
                        if position is not None:
                            logical_columns[definition.name] = position + 1
                            break
                data_end_row = _table_data_end_row(
                    table,
                    table.header_row,
                    sheet,
                    formula_workbook[sheet_name],
                    logical_columns,
                    value_resolution,
                    formula_resolution,
                )
                if data_end_row is None:
                    errors.append(
                        ValidationEvent(
                            rule_code="table_resolution",
                            sheet_name=sheet_name,
                            description=(
                                f"Table '{table.name}' data boundary "
                                "could not be resolved"
                            ),
                        )
                    )
                required_positions = []
                missing_columns = []
                for logical_name in table.header_policy.required_columns:
                    definition = definitions.get(logical_name)
                    names = [logical_name]
                    if definition is not None:
                        names.extend(definition.aliases)
                    position = next(
                        (
                            normalised_headers[_normalise_header(name)]
                            for name in names
                            if _normalise_header(name) in normalised_headers
                        ),
                        None,
                    )
                    if position is None:
                        missing_columns.append(logical_name)
                    else:
                        required_positions.append(position)
                if table.header_policy.match_mode == "exact_order":
                    if required_positions != list(range(len(required_positions))):
                        missing_columns.append("required columns are out of order")
                elif (
                    table.header_policy.match_mode == "contains_in_order"
                    and required_positions != sorted(required_positions)
                ):
                    missing_columns.append("required columns are out of order")
                header_resolution_failed = bool(
                    missing_columns or physical_header_issues
                )
                boundary_resolution_failed = (
                    data_end_row is None
                    or (
                        table.data_boundary is not None
                        and table.data_boundary.mode == "last_non_empty_row"
                        and any(
                            column not in logical_columns
                            for column in (table.data_boundary.columns or [])
                        )
                    )
                )
                if missing_columns and header_rule_enabled:
                    event = ValidationEvent(
                        rule_code="column_header",
                        sheet_name=sheet_name,
                        actual_value=headers,
                        expected_value=table.header_policy.required_columns,
                        description=(
                            f"Table '{table.name}' header validation failed: "
                            f"{', '.join(missing_columns)}"
                        ),
                    )
                    if table.header_policy.missing_column_action == "warning":
                        warnings.append(event)
                    else:
                        errors.append(event)
                known_headers = {
                    _normalise_header(name)
                    for definition in table.column_definitions
                    for name in [definition.name, *definition.aliases]
                }
                extra_headers = [
                    header
                    for header in headers
                    if header is not None
                    and _normalise_header(header) not in known_headers
                ]
                if (
                    extra_headers
                    and header_rule_enabled
                    and table.header_policy.extra_column_action != "allowed"
                ):
                    for header in extra_headers:
                        event = ValidationEvent(
                            rule_code="column_header",
                            sheet_name=sheet_name,
                            actual_value=header,
                            expected_value="known table header",
                            description=(
                                f"Extra table header found in '{table.name}': "
                                f"{header}"
                            ),
                        )
                        if table.header_policy.extra_column_action == "warning":
                            warnings.append(event)
                        else:
                            errors.append(event)
                if table.data_presence == "require_one_usable_row":
                    if not data_presence_rule_enabled:
                        not_run.append(
                            ValidationEvent(
                                rule_code="non_empty_data",
                                status="NOT_RUN",
                                reason="RULE_DISABLED",
                                description=(
                                    f"Required table '{table.name}' data "
                                    "presence check is disabled by configuration"
                                ),
                            )
                        )
                    elif boundary_resolution_failed:
                        not_run.append(
                            ValidationEvent(
                                rule_code="non_empty_data",
                                status="NOT_RUN",
                                reason="DATA_BOUNDARY_RESOLUTION_FAILED",
                                description=(
                                    f"Table '{table.name}' data boundary "
                                    "could not be resolved"
                                ),
                            )
                        )
                    else:
                        formula_sheet = formula_workbook[sheet_name]
                        has_data_row = False
                        max_row = data_end_row
                        max_column = max(sheet.max_column, formula_sheet.max_column)
                        for row_number in range(table.header_row + 1, max_row + 1):
                            if any(
                                sheet.cell(row_number, column).value is not None
                                or formula_sheet.cell(row_number, column).value is not None
                                for column in range(1, max_column + 1)
                            ):
                                has_data_row = True
                                break
                        if not has_data_row:
                            errors.append(
                                ValidationEvent(
                                    rule_code="non_empty_data",
                                    sheet_name=sheet_name,
                                    expected_value="at least one usable data row",
                                    description=(
                                        f"Required table '{table.name}' has no "
                                        "usable data rows"
                                    ),
                                )
                            )
                formula_sheet = formula_workbook[sheet_name]
                max_column = max(sheet.max_column, formula_sheet.max_column)
                if table.empty_row_rules and not empty_row_rule_enabled:
                    not_run.append(
                        ValidationEvent(
                            rule_code="empty_row_pattern",
                            status="NOT_RUN",
                            reason="RULE_DISABLED",
                            description=(
                                f"Empty row checks for table '{table.name}' "
                                "are disabled by configuration"
                            ),
                        )
                    )
                for empty_row_rule in (
                    table.empty_row_rules if empty_row_rule_enabled else []
                ):
                    excluded_columns = {
                        column_index_from_string(column)
                        for column in empty_row_rule.excluded_columns
                    }
                    for row_number in empty_row_rule.rows:
                        for column_number in range(1, max_column + 1):
                            if column_number in excluded_columns:
                                continue
                            cell = sheet.cell(row_number, column_number)
                            if _is_blank_value(
                                cell.value,
                                empty_row_rule.blank_policy,
                                resolved_config.comparison.null_tokens,
                            ):
                                continue
                            errors.append(
                                ValidationEvent(
                                    rule_code="empty_row_pattern",
                                    sheet_name=sheet_name,
                                    cell_reference=cell.coordinate,
                                    row_number=row_number,
                                    actual_value=cell.value,
                                    expected_value="blank",
                                    description=(
                                        f"Configured empty row {row_number} "
                                        "contains a value"
                                    ),
                                )
                            )
                max_row = data_end_row
                if boundary_resolution_failed:
                    for validation in table.column_validations:
                        column_rule_codes = []
                        if validation.required:
                            column_rule_codes.append("missing_value")
                        if validation.unique:
                            column_rule_codes.append("duplicate_value")
                        if validation.allowed_values is not None:
                            column_rule_codes.append("allowed_values")
                        if validation.forbidden_values is not None:
                            column_rule_codes.append("forbidden_values")
                        for rule_code in column_rule_codes:
                            not_run.append(
                                ValidationEvent(
                                    rule_code=rule_code,
                                    status="NOT_RUN",
                                    reason="DATA_BOUNDARY_RESOLUTION_FAILED",
                                    description=(
                                        f"Column rule for '{validation.column}' "
                                        "requires a resolved data boundary"
                                    ),
                                )
                            )
                    continue
                if header_resolution_failed:
                    for validation in table.column_validations:
                        column_rule_codes = []
                        if validation.required:
                            column_rule_codes.append("missing_value")
                        if validation.unique:
                            column_rule_codes.append("duplicate_value")
                        if validation.allowed_values is not None:
                            column_rule_codes.append("allowed_values")
                        if validation.forbidden_values is not None:
                            column_rule_codes.append("forbidden_values")
                        for rule_code in column_rule_codes:
                            disabled = (
                                enabled_rules is not None
                                and rule_code not in enabled_rules
                            )
                            not_run.append(
                                ValidationEvent(
                                    rule_code=rule_code,
                                    status="NOT_RUN",
                                    reason=(
                                        "RULE_DISABLED"
                                        if disabled
                                        else "TABLE_HEADER_RESOLUTION_FAILED"
                                    ),
                                    severity=validation.severity
                                    or resolved_config.rule_severity.get(rule_code),
                                    description=(
                                        f"Column rule for '{validation.column}' "
                                        + (
                                            "is disabled by configuration"
                                            if disabled
                                            else "requires resolved table headers"
                                        )
                                    ),
                                )
                            )
                    continue
                for validation in table.column_validations:
                    for rule_code, enabled in (
                        ("missing_value", validation.required or validation.null_policy == "error"),
                        ("duplicate_value", validation.unique),
                        ("allowed_values", validation.allowed_values is not None),
                        ("forbidden_values", validation.forbidden_values is not None),
                    ):
                        if enabled:
                            processed_severities.setdefault(
                                rule_code,
                                validation.severity
                                or resolved_config.rule_severity.get(rule_code),
                            )
                    column_number = logical_columns.get(validation.column)
                    if column_number is None:
                        continue
                    column_definition = definitions.get(validation.column)
                    comparison = (
                        column_definition.comparison
                        if column_definition is not None
                        and column_definition.comparison is not None
                        else resolved_config.comparison
                    )
                    column_rule_codes = []
                    if validation.required:
                        column_rule_codes.append("missing_value")
                    if validation.unique:
                        column_rule_codes.append("duplicate_value")
                    if validation.allowed_values is not None:
                        column_rule_codes.append("allowed_values")
                    if validation.forbidden_values is not None:
                        column_rule_codes.append("forbidden_values")
                    disabled_column_rules = [
                        rule_code
                        for rule_code in column_rule_codes
                        if (
                            resolved_config.enabled_rules is not None
                            and rule_code not in resolved_config.enabled_rules
                        )
                    ]
                    if disabled_column_rules:
                        for rule_code in disabled_column_rules:
                            not_run.append(
                                ValidationEvent(
                                    rule_code=rule_code,
                                    status="NOT_RUN",
                                    reason="RULE_DISABLED",
                                    severity=validation.severity
                                    or resolved_config.rule_severity.get(rule_code),
                                    description=(
                                        f"Column rule for '{validation.column}' "
                                        "is disabled by configuration"
                                    ),
                                )
                            )
                        if len(disabled_column_rules) == len(column_rule_codes):
                            continue
                    missing_enabled = "missing_value" not in disabled_column_rules
                    duplicate_enabled = "duplicate_value" not in disabled_column_rules
                    allowed_enabled = "allowed_values" not in disabled_column_rules
                    forbidden_enabled = "forbidden_values" not in disabled_column_rules
                    seen_values = {}
                    for row_number in range(table.header_row + 1, max_row + 1):
                        cell = sheet.cell(row_number, column_number)
                        actual_value = cell.value
                        for rule_code, enabled in (
                            (
                                "missing_value",
                                missing_enabled
                                and (
                                    validation.required
                                    or validation.null_policy == "error"
                                ),
                            ),
                            ("duplicate_value", duplicate_enabled and validation.unique),
                            (
                                "allowed_values",
                                allowed_enabled and validation.allowed_values is not None,
                            ),
                            (
                                "forbidden_values",
                                forbidden_enabled and validation.forbidden_values is not None,
                            ),
                        ):
                            if enabled:
                                record_processed(rule_code)
                        normalised_value = _normalise_comparison_value(
                            actual_value,
                            comparison,
                        )
                        is_null = normalised_value is None
                        if (
                            is_null
                            and missing_enabled
                            and (
                                validation.required
                                or validation.null_policy == "error"
                            )
                        ):
                            errors.append(
                                ValidationEvent(
                                    rule_code="missing_value",
                                    severity=validation.severity
                                    or resolved_config.rule_severity.get(
                                        "missing_value"
                                    ),
                                    sheet_name=sheet_name,
                                    cell_reference=cell.coordinate,
                                    row_number=row_number,
                                    actual_value=actual_value,
                                    expected_value="non-null value",
                                    description=(
                                        f"Required value is missing for column "
                                        f"'{validation.column}'"
                                    ),
                                )
                            )
                        if is_null and validation.null_policy in {
                            "allow",
                            "ignore",
                        }:
                            continue
                        if validation.unique and duplicate_enabled:
                            value_key = repr(normalised_value)
                            if value_key in seen_values:
                                errors.append(
                                    ValidationEvent(
                                        rule_code="duplicate_value",
                                        severity=validation.severity
                                        or resolved_config.rule_severity.get(
                                            "duplicate_value"
                                        ),
                                        sheet_name=sheet_name,
                                        cell_reference=cell.coordinate,
                                        row_number=row_number,
                                        actual_value=actual_value,
                                        expected_value="unique value",
                                        description=(
                                            f"Duplicate value found for column "
                                            f"'{validation.column}'"
                                        ),
                                    )
                                )
                            else:
                                seen_values[value_key] = cell.coordinate
                        if (
                            validation.allowed_values is not None
                            and allowed_enabled
                            and normalised_value
                            not in {
                                    _normalise_comparison_value(
                                        value,
                                        comparison,
                                    )
                                for value in validation.allowed_values
                            }
                        ):
                            errors.append(
                                ValidationEvent(
                                    rule_code="allowed_values",
                                    severity=validation.severity
                                    or resolved_config.rule_severity.get(
                                        "allowed_values"
                                    ),
                                    sheet_name=sheet_name,
                                    cell_reference=cell.coordinate,
                                    row_number=row_number,
                                    actual_value=actual_value,
                                    expected_value=validation.allowed_values,
                                    description=(
                                        f"Value is not allowed for column "
                                        f"'{validation.column}'"
                                    ),
                                )
                            )
                        if (
                            validation.forbidden_values is not None
                            and forbidden_enabled
                            and normalised_value
                            in {
                                    _normalise_comparison_value(
                                        value,
                                        comparison,
                                    )
                                for value in validation.forbidden_values
                            }
                        ):
                            errors.append(
                                ValidationEvent(
                                    rule_code="forbidden_values",
                                    severity=validation.severity
                                    or resolved_config.rule_severity.get(
                                        "forbidden_values"
                                    ),
                                    sheet_name=sheet_name,
                                    cell_reference=cell.coordinate,
                                    row_number=row_number,
                                    actual_value=actual_value,
                                    expected_value=validation.forbidden_values,
                                    description=(
                                        f"Value is forbidden for column "
                                        f"'{validation.column}'"
                                    ),
                                )
                            )
        dataframe_errors, dataframe_not_run, dataframe_counts, dataframe_cache = (
            _run_dataframe_checks(
                candidate_path,
                value_workbook,
                formula_workbook,
                resolved_config,
                value_resolution,
                formula_resolution,
            )
        )
        errors.extend(dataframe_errors)
        not_run.extend(dataframe_not_run)
        for rule_code, count in dataframe_counts.items():
            processed_counts[rule_code] = (
                processed_counts.get(rule_code, 0) + count
            )
        cross_errors, cross_not_run = _run_cross_table_checks(
            resolved_config,
            dataframe_cache,
        )
        errors.extend(cross_errors)
        not_run.extend(cross_not_run)
        rule_outcomes = {}
        workbook_checks = _ordered_workbook_checks(
            resolved_config.workbook_checks
        )
        for check_index, check in enumerate(workbook_checks):
            if any(
                rule_outcomes.get(dependency) != "passed"
                for dependency in check.depends_on
            ):
                failed_dependencies = [
                    dependency
                    for dependency in check.depends_on
                    if rule_outcomes.get(dependency) != "passed"
                ]
                dependency_reason = (
                    "DEPENDENCY_FAILED"
                    if any(
                        rule_outcomes.get(dependency) == "failed"
                        for dependency in failed_dependencies
                    )
                    else "DEPENDENCY_NOT_RUN"
                )
                record_processed(check.rule_code, 0)
                not_run.append(
                    ValidationEvent(
                        rule_code=check.rule_code,
                        status="NOT_RUN",
                        reason=dependency_reason,
                        severity=check.severity
                        or resolved_config.rule_severity.get(check.rule_code),
                        expected_value=failed_dependencies,
                        description=(
                            f"Rule depends on checks that did not pass: "
                            f"{', '.join(failed_dependencies)}"
                        ),
                    )
                )
                rule_outcomes[check.rule_code] = "not_run"
                continue
            if (
                not check.enabled
                or (
                    resolved_config.enabled_rules is not None
                    and check.rule_code not in resolved_config.enabled_rules
                )
            ):
                record_processed(check.rule_code, 0)
                not_run.append(
                    ValidationEvent(
                        rule_code=check.rule_code,
                        status="NOT_RUN",
                        reason="RULE_DISABLED",
                        severity=check.severity
                        or resolved_config.rule_severity.get(check.rule_code),
                        description="Rule is disabled by configuration",
                    )
                )
                rule_outcomes[check.rule_code] = "not_run"
                continue
            if check.rule_code != "formula_error":
                record_processed(check.rule_code)
            custom_rule = get_registered_rule(check.rule_code)
            if custom_rule is not None:
                context = RuleContext(
                    rule_code=check.rule_code,
                    candidate_path=str(candidate_path),
                    reference_path=(
                        str(reference_path)
                        if reference_path is not None
                        else None
                    ),
                    config=resolved_config,
                )
                events = validate_custom_result(custom_rule(context))
                for event in events:
                    if event.status == "NOT_RUN":
                        not_run.append(event)
                    elif event.severity == "warning":
                        warnings.append(event)
                    else:
                        errors.append(event)
                if any(event.rule_code == check.rule_code for event in errors):
                    rule_outcomes[check.rule_code] = "failed"
                elif any(event.rule_code == check.rule_code for event in not_run):
                    rule_outcomes[check.rule_code] = "not_run"
                else:
                    rule_outcomes[check.rule_code] = "passed"
                continue
            if check.rule_code == "missing_reference_sheet":
                if reference_workbook is None:
                    not_run.append(
                        ValidationEvent(
                            rule_code="missing_reference_sheet",
                            status="NOT_RUN",
                            reason="REFERENCE_WORKBOOK_REQUIRED",
                            description=(
                                "Reference workbook is required for this check"
                            ),
                        )
                    )
                    rule_outcomes[check.rule_code] = "not_run"
                    continue
                candidate_names = set(formula_workbook.sheetnames)
                for sheet_name in reference_workbook.sheetnames:
                    if sheet_name not in candidate_names:
                        errors.append(
                            ValidationEvent(
                                rule_code="missing_reference_sheet",
                                sheet_name=sheet_name,
                                description=(
                                    f"Reference sheet missing from candidate: "
                                    f"{sheet_name}"
                                ),
                            )
                        )
                rule_outcomes[check.rule_code] = (
                    "failed"
                    if any(event.rule_code == check.rule_code for event in errors)
                    else "passed"
                )
                continue
            if check.rule_code == "sheet_structure":
                if reference_workbook is None:
                    not_run.append(
                        ValidationEvent(
                            rule_code="sheet_structure",
                            status="NOT_RUN",
                            reason="REFERENCE_WORKBOOK_REQUIRED",
                            description=(
                                "Reference workbook is required for this check"
                            ),
                        )
                    )
                    rule_outcomes[check.rule_code] = "not_run"
                    continue
                reference_sheets = {
                    sheet.title: sheet for sheet in reference_workbook.worksheets
                }
                reference_formula_sheets = {
                    sheet.title: sheet
                    for sheet in reference_formula_workbook.worksheets
                }
                candidate_formula_sheets = {
                    sheet.title: sheet
                    for sheet in formula_workbook.worksheets
                }
                for candidate_sheet in value_workbook.worksheets:
                    reference_sheet = reference_sheets.get(candidate_sheet.title)
                    if reference_sheet is None:
                        continue
                    actual_shape = value_resolution.shape(
                        candidate_sheet,
                        candidate_formula_sheets[candidate_sheet.title],
                    )
                    expected_shape = reference_value_resolution.shape(
                        reference_sheet,
                        reference_formula_sheets[reference_sheet.title],
                    )
                    if actual_shape == expected_shape:
                        continue
                    errors.append(
                        ValidationEvent(
                            rule_code="sheet_structure",
                            sheet_name=candidate_sheet.title,
                            actual_value=actual_shape,
                            expected_value=expected_shape,
                            description=(
                                "Candidate used area differs from reference"
                            ),
                        )
                    )
                rule_outcomes[check.rule_code] = (
                    "failed"
                    if any(event.rule_code == check.rule_code for event in errors)
                    else "passed"
                )
                continue
            if check.rule_code == "formula_difference":
                whitespace_policy = (check.options or {}).get(
                    "whitespace_policy",
                    "normalised",
                )
                if whitespace_policy not in {"exact", "normalised"}:
                    raise ConfigurationError(
                        "must be 'exact' or 'normalised'",
                        field_path=(
                            f"workbook_checks[{check_index}]"
                            ".options.whitespace_policy"
                        ),
                    )
                if reference_formula_workbook is None:
                    not_run.append(
                        ValidationEvent(
                            rule_code="formula_difference",
                            status="NOT_RUN",
                            reason="REFERENCE_WORKBOOK_REQUIRED",
                            description=(
                                "Reference workbook is required for this check"
                            ),
                        )
                    )
                    rule_outcomes[check.rule_code] = "not_run"
                    continue
                reference_sheets = {
                    sheet.title: sheet for sheet in reference_formula_workbook.worksheets
                }
                for candidate_sheet in formula_workbook.worksheets:
                    reference_sheet = reference_sheets.get(candidate_sheet.title)
                    if reference_sheet is None:
                        continue
                    for row_number, column_number in _comparison_coordinates(
                        candidate_sheet,
                        reference_sheet,
                        formula_resolution,
                        reference_formula_resolution,
                    ):
                            candidate_cell = candidate_sheet.cell(
                                row=row_number,
                                column=column_number,
                            )
                            reference_cell = reference_sheet.cell(
                                row=row_number,
                                column=column_number,
                            )
                            candidate_is_formula = candidate_cell.data_type == "f"
                            reference_is_formula = reference_cell.data_type == "f"
                            if not (candidate_is_formula or reference_is_formula):
                                continue
                            candidate_formula = _normalise_formula(
                                candidate_cell.value,
                                whitespace_policy,
                            )
                            reference_formula = _normalise_formula(
                                reference_cell.value,
                                whitespace_policy,
                            )
                            if candidate_formula == reference_formula:
                                continue
                            errors.append(
                                ValidationEvent(
                                    rule_code="formula_difference",
                                    sheet_name=candidate_sheet.title,
                                    cell_reference=candidate_cell.coordinate,
                                    actual_value=candidate_formula,
                                    expected_value=reference_formula,
                                    description=(
                                        "Candidate formula differs from reference"
                                    ),
                                )
                            )
                rule_outcomes[check.rule_code] = (
                    "failed"
                    if any(event.rule_code == check.rule_code for event in errors)
                    else "passed"
                )
                continue
            if check.rule_code != "formula_error":
                rule_outcomes[check.rule_code] = "passed"
                continue
            processed_counts.setdefault("formula_error", 0)
            tokens = (check.options or {}).get("error_tokens", [])
            severity = check.severity or resolved_config.rule_severity.get(
                "formula_error"
            )
            missing_cache_action = (
                resolved_config.runtime.missing_formula_cache_action
            )
            missing_cache_cells = []
            for formula_sheet in formula_workbook.worksheets:
                value_sheet = value_workbook[formula_sheet.title]
                for formula_cell in formula_resolution.cells(formula_sheet):
                    value_cell = value_sheet[formula_cell.coordinate]
                    if (
                        formula_cell.data_type == "f"
                        and value_cell.value is None
                        and value_cell.data_type not in {"s", "str"}
                    ):
                        missing_cache_cells.append(
                            (
                                formula_sheet.title,
                                formula_cell.coordinate,
                                formula_cell.row,
                            )
                        )
            if missing_cache_cells:
                status = (
                    "NOT_RUN"
                    if missing_cache_action == "not_run"
                    else "FAILED"
                )
                for sheet_name, cell_reference, row_number in missing_cache_cells:
                    event = ValidationEvent(
                        rule_code="formula_error",
                        status=status,
                        reason="FORMULA_RESULTS_UNAVAILABLE",
                        severity=severity,
                        sheet_name=sheet_name,
                        cell_reference=cell_reference,
                        row_number=row_number,
                        description=(
                            "Formula result is unavailable because the "
                            "workbook has no cached value"
                        ),
                    )
                    if status == "NOT_RUN":
                        not_run.append(event)
                    elif severity == "warning":
                        warnings.append(event)
                    else:
                        errors.append(event)
                if status == "NOT_RUN":
                    rule_outcomes[check.rule_code] = "not_run"
                    continue
            for sheet in value_workbook.worksheets:
                for cell in value_resolution.cells(sheet):
                    record_processed("formula_error")
                    if cell.data_type == "e" or cell.value in tokens:
                        event = ValidationEvent(
                            rule_code="formula_error",
                            reason="FORMULA_ERROR_VALUE",
                            severity=severity,
                            sheet_name=sheet.title,
                            cell_reference=cell.coordinate,
                            row_number=cell.row,
                            actual_value=cell.value,
                            expected_value=None,
                            description=(
                                f"Excel error value found: {cell.value}"
                            ),
                        )
                        if severity == "warning":
                            warnings.append(event)
                        else:
                            errors.append(event)
            rule_outcomes[check.rule_code] = (
                "failed"
                if any(event.rule_code == check.rule_code for event in errors)
                else "passed"
            )
        finalize_processed_counts()
        return ValidationResult(
            **result_metadata,
            processed_counts=processed_counts,
            processed_severities=processed_severities,
            errors=errors,
            warnings=warnings,
            not_run=not_run,
            diagnostics=diagnostics,
        )
    finally:
        formula_workbook.close()
        value_workbook.close()
        if reference_workbook is not None:
            reference_workbook.close()
        if reference_formula_workbook is not None:
            reference_formula_workbook.close()


def _run_dataframe_checks(
    candidate_path,
    value_workbook,
    formula_workbook,
    config,
    value_resolution,
    formula_resolution,
):
    errors = []
    not_run = []
    processed_counts = {}
    dataframe_cache = {}

    def skip_dataframe_checks(table, reason, sheet_name=None):
        for check in table.dataframe_checks:
            not_run.append(
                ValidationEvent(
                    rule_code=check.rule_code,
                    status="NOT_RUN",
                    reason=reason,
                    sheet_name=sheet_name,
                    description=(
                        f"DataFrame check for '{check.column}' could not run: "
                        f"{reason}"
                    ),
                )
            )

    for table in config.tables:
        if table.load_policy is None or not table.load_policy.enabled:
            continue
        if config.enabled_rules is not None and "pandas_load" not in config.enabled_rules:
            not_run.append(
                ValidationEvent(
                    rule_code="pandas_load",
                    status="NOT_RUN",
                    reason="RULE_DISABLED",
                    description=f"Pandas loading for table '{table.name}' is disabled",
                )
            )
            skip_dataframe_checks(table, "PANDAS_LOAD_DISABLED")
            continue
        if table.sheet_selector is None or table.header_row is None:
            not_run.append(
                ValidationEvent(
                    rule_code="pandas_load",
                    status="NOT_RUN",
                    reason="TABLE_CONFIGURATION_INCOMPLETE",
                    description=f"Table '{table.name}' is not configured for pandas loading",
                )
            )
            skip_dataframe_checks(table, "TABLE_CONFIGURATION_INCOMPLETE")
            continue
        sheet_matches = _match_sheets(
            value_workbook.sheetnames,
            table.sheet_selector,
        )
        if not sheet_matches:
            not_run.append(
                ValidationEvent(
                    rule_code="pandas_load",
                    status="NOT_RUN",
                    reason="TABLE_RESOLUTION_FAILED",
                    description=f"Table '{table.name}' did not resolve",
                )
            )
            skip_dataframe_checks(table, "TABLE_RESOLUTION_FAILED")
            continue
        for sheet_name in sheet_matches:
            sheet = value_workbook[sheet_name]
            formula_sheet = formula_workbook[sheet_name]
            headers = value_resolution.headers(sheet, table.header_row, table)
            while headers and headers[-1] is None:
                headers.pop()
            definitions = {
                definition.name: definition
                for definition in table.column_definitions
            }
            normalised_headers = {
                _normalise_header(header): index + 1
                for index, header in enumerate(headers)
                if header is not None
            }
            logical_columns = {
                name: next(
                    (
                        normalised_headers[_normalise_header(candidate)]
                        for candidate in [name, *definitions.get(name, ColumnDefinition(name=name)).aliases]
                        if _normalise_header(candidate) in normalised_headers
                    ),
                    None,
                )
                for name in definitions
            }
            required_columns = (
                table.header_policy.required_columns
                if table.header_policy is not None
                else []
            )
            if any(column not in logical_columns for column in required_columns):
                not_run.append(
                    ValidationEvent(
                        rule_code="pandas_load",
                        status="NOT_RUN",
                        reason="TABLE_HEADER_RESOLUTION_FAILED",
                        sheet_name=sheet_name,
                        description=f"Table '{table.name}' headers could not be resolved",
                    )
                )
                skip_dataframe_checks(table, "TABLE_HEADER_RESOLUTION_FAILED", sheet_name)
                continue
            data_end_row = _table_data_end_row(
                table,
                table.header_row,
                sheet,
                formula_sheet,
                logical_columns,
                value_resolution,
                formula_resolution,
            )
            if data_end_row is None:
                not_run.append(
                    ValidationEvent(
                        rule_code="pandas_load",
                        status="NOT_RUN",
                        reason="DATA_BOUNDARY_RESOLUTION_FAILED",
                        sheet_name=sheet_name,
                        description=f"Table '{table.name}' data boundary could not be resolved",
                    )
                )
                skip_dataframe_checks(table, "DATA_BOUNDARY_RESOLUTION_FAILED", sheet_name)
                continue
            try:
                dataframe = pd.read_excel(
                    candidate_path,
                    sheet_name=sheet_name,
                    header=table.header_row - 1,
                    usecols=f"A:{get_column_letter(len(headers))}",
                    nrows=max(0, data_end_row - table.header_row),
                    dtype=None if table.load_policy.infer_types else object,
                    keep_default_na=not table.load_policy.preserve_empty_values,
                )
            except Exception as error:
                errors.append(
                    ValidationEvent(
                        rule_code="pandas_load",
                        sheet_name=sheet_name,
                        actual_value=str(error),
                        description=f"Could not load table '{table.name}' into pandas: {error}",
                    )
                )
                continue
            processed_counts["pandas_load"] = processed_counts.get("pandas_load", 0) + 1
            dataframe_cache[table.name] = dataframe
            for check in table.dataframe_checks:
                enabled = (
                    check.enabled
                    and (
                        config.enabled_rules is None
                        or check.rule_code in config.enabled_rules
                    )
                )
                if not enabled:
                    not_run.append(
                        ValidationEvent(
                            rule_code=check.rule_code,
                            status="NOT_RUN",
                            reason="RULE_DISABLED",
                            sheet_name=sheet_name,
                            description=f"DataFrame check for '{check.column}' is disabled",
                        )
                    )
                    continue
                if check.column not in dataframe.columns:
                    errors.append(
                        ValidationEvent(
                            rule_code=check.rule_code,
                            sheet_name=sheet_name,
                            description=f"DataFrame column not found: {check.column}",
                        )
                    )
                    continue
                column_number = logical_columns.get(check.column)
                if column_number is None:
                    continue
                for dataframe_row, value in dataframe[check.column].items():
                    if value is None or (isinstance(value, float) and pd.isna(value)):
                        continue
                    if len(str(value)) <= check.max_length:
                        continue
                    excel_row = table.header_row + 1 + int(dataframe_row)
                    errors.append(
                        ValidationEvent(
                            rule_code=check.rule_code,
                            severity=check.severity
                            or config.rule_severity.get(check.rule_code),
                            sheet_name=sheet_name,
                            cell_reference=f"{get_column_letter(column_number)}{excel_row}",
                            row_number=excel_row,
                            column_number=column_number,
                            actual_value=value,
                            expected_value=check.max_length,
                            description=(
                                f"Value in column '{check.column}' exceeds "
                                f"the maximum length of {check.max_length}"
                            ),
                        )
                    )
    return errors, not_run, processed_counts, dataframe_cache


def _run_cross_table_checks(config, dataframe_cache):
    errors = []
    not_run = []
    for check in config.cross_table_checks:
        if (
            not check.enabled
            or (
                config.enabled_rules is not None
                and check.rule_code not in config.enabled_rules
            )
        ):
            not_run.append(
                ValidationEvent(
                    rule_code=check.rule_code,
                    status="NOT_RUN",
                    reason="RULE_DISABLED",
                    description=f"Cross-table check '{check.name}' is disabled",
                )
            )
            continue
        source = dataframe_cache.get(check.source_table)
        reference = dataframe_cache.get(check.reference_table)
        if source is None or reference is None:
            not_run.append(
                ValidationEvent(
                    rule_code=check.rule_code,
                    status="NOT_RUN",
                    reason="PANDAS_LOAD_FAILED",
                    description=(
                        f"Cross-table check '{check.name}' requires loaded "
                        f"tables '{check.source_table}' and '{check.reference_table}'"
                    ),
                )
            )
            continue
        if check.source_column not in source.columns:
            errors.append(
                ValidationEvent(
                    rule_code=check.rule_code,
                    severity=check.severity,
                    description=f"DataFrame column not found: {check.source_column}",
                )
            )
            continue
        if check.reference_column not in reference.columns:
            errors.append(
                ValidationEvent(
                    rule_code=check.rule_code,
                    severity=check.severity,
                    description=f"DataFrame column not found: {check.reference_column}",
                )
            )
            continue
        source_config = next(
            table
            for table in config.tables
            if table.name == check.source_table
        )
        source_sheet_name = (
            source_config.sheet_selector.value
            if source_config.sheet_selector is not None
            and source_config.sheet_selector.mode == "exact"
            else None
        )
        source_column_number = source.columns.get_loc(check.source_column) + 1
        allowed = {
            value
            for value in reference[check.reference_column].tolist()
            if value is not None and not (isinstance(value, float) and pd.isna(value))
        }
        for row_index, value in source[check.source_column].items():
            if value is None or (isinstance(value, float) and pd.isna(value)):
                continue
            if value in allowed:
                continue
            errors.append(
                ValidationEvent(
                    rule_code=check.rule_code,
                    severity=check.severity,
                    actual_value=value,
                    expected_value=check.reference_table,
                    sheet_name=source_sheet_name,
                    cell_reference=(
                        f"{get_column_letter(source_column_number)}"
                        f"{source_config.header_row + 1 + int(row_index)}"
                    ),
                    row_number=source_config.header_row + 1 + int(row_index),
                    column_number=source_column_number,
                    description=(
                        f"Value in '{check.source_column}' was not found in "
                        f"'{check.reference_table}.{check.reference_column}'"
                    ),
                )
            )
    return errors, not_run


def _match_sheets(sheet_names, selector):
    if selector.mode == "exact":
        if selector.case_sensitive:
            return [name for name in sheet_names if name == selector.value]
        expected = selector.value.casefold()
        return [name for name in sheet_names if name.casefold() == expected]
    if selector.mode == "regex":
        try:
            pattern = re.compile(
                selector.value,
                0 if selector.case_sensitive else re.IGNORECASE,
            )
        except re.error as error:
            raise ConfigurationError(
                f"invalid sheet selector regex: {error}"
            ) from error
        return [
            name
            for name in sheet_names
            if pattern.fullmatch(name) is not None
        ]
    return []


def _select_sheet_matches(matches, action, selector, errors):
    if action == "first":
        return matches[:1]
    if action == "error_on_multiple" and len(matches) > 1:
        errors.append(
            ValidationEvent(
                rule_code="sheet_selector",
                actual_value=len(matches),
                expected_value=1,
                description=(
                    f"Sheet selector matched multiple sheets: {selector.value}"
                ),
            )
        )
    return matches


def _normalise_formula(value, whitespace_policy="normalised"):
    text = getattr(value, "text", None)
    if text is None:
        return _normalise_formula_text(value, whitespace_policy)
    return {
        "text": _normalise_formula_text(text, whitespace_policy),
        "ref": getattr(value, "ref", None),
    }


def _normalise_formula_text(value, whitespace_policy):
    if not isinstance(value, str):
        return value
    if whitespace_policy == "exact":
        return value
    text = value.strip()
    if text.startswith("="):
        return "=" + text[1:].lstrip()
    return text


def _values_equal(actual, expected, mode, options=None):
    options = options or {}
    if mode == "exact":
        return actual == expected
    if mode in {"normalised", "case_insensitive"}:
        if isinstance(actual, str) and isinstance(expected, str):
            actual_text = actual.strip()
            expected_text = expected.strip()
            if mode == "case_insensitive":
                return actual_text.casefold() == expected_text.casefold()
            return actual_text == expected_text
    if mode == "numeric":
        actual_number = _numeric_value(actual, options)
        expected_number = _numeric_value(expected, options)
        return (
            actual_number is not None
            and expected_number is not None
            and actual_number == expected_number
        )
    if mode == "date":
        actual_date = _date_value(actual, options)
        expected_date = _date_value(expected, options)
        return (
            actual_date is not None
            and expected_date is not None
            and actual_date == expected_date
        )
    return actual == expected


def _normalise_comparison_value(value, comparison):
    if value is None:
        return None
    if not isinstance(value, str):
        if value in comparison.null_tokens:
            return None
        return value
    text = _normalise_text(value, comparison)
    if comparison.empty_string_is_null and text == "":
        return None
    null_tokens = {
        _normalise_text(token, comparison)
        for token in comparison.null_tokens
        if isinstance(token, str)
    }
    if text in null_tokens:
        return None
    return text


def _normalise_text(value, comparison):
    text = value
    if comparison.trim_whitespace:
        text = text.strip()
    if comparison.collapse_internal_whitespace:
        text = re.sub(r"\s+", " ", text)
    if not comparison.accent_sensitive:
        text = "".join(
            character
            for character in unicodedata.normalize("NFD", text)
            if not unicodedata.combining(character)
        )
    if not comparison.case_sensitive:
        text = text.casefold()
    return text


def _numeric_value(value, options):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        number = Decimal(str(value))
        return number if number.is_finite() else None
    if not options.get("allow_numeric_strings") or not isinstance(value, str):
        return None
    text = value.strip().replace(
        options["thousands_separator"],
        "",
    )
    text = text.replace(options["decimal_separator"], ".")
    try:
        number = Decimal(text)
        return number if number.is_finite() else None
    except (InvalidOperation, ValueError):
        return None


def _date_value(value, options):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None
    for accepted_format in options["accepted_formats"]:
        try:
            return datetime.strptime(value.strip(), accepted_format).date()
        except ValueError:
            continue
    return None


def _detected_features(workbook, candidate_path):
    if candidate_path.suffix.lower() == ".xlsm":
        yield "macros", None, None, None, None
    for feature_name, attribute in (
        ("external_links", "_external_links"),
        ("named_ranges", "defined_names"),
    ):
        if feature_name == "external_links":
            targets = _external_link_targets(workbook, candidate_path)
            if targets:
                for target in targets:
                    yield feature_name, None, None, target, None
                continue
        if _package_feature_present(candidate_path, feature_name):
            yield feature_name, None, None, None, None
            continue
        try:
            if getattr(workbook, attribute):
                yield feature_name, None, None, None, None
        except Exception as error:
            yield feature_name, None, None, None, str(error)
    package_features = tuple(
        feature_name
        for feature_name in ("charts", "pivot_tables", "merged_cells")
        if _package_feature_present(candidate_path, feature_name)
    )
    for feature_name in package_features:
        if getattr(workbook, "read_only", False):
            yield (
                feature_name,
                None,
                None,
                None,
                "openpyxl read-only mode does not expose this feature",
            )
        elif feature_name != "merged_cells":
            yield feature_name, None, None, None, None
    for sheet in workbook.worksheets:
        for feature_name, attribute in (
            ("charts", "_charts"),
            ("pivot_tables", "_pivots"),
            ("merged_cells", "merged_cells"),
        ):
            if (
                feature_name in package_features
                and feature_name != "merged_cells"
            ):
                continue
            try:
                feature = getattr(sheet, attribute, None)
                if feature_name == "merged_cells":
                    present = feature is not None and bool(list(feature.ranges))
                else:
                    present = bool(feature)
                if present:
                    if feature_name == "merged_cells":
                        for merged_range in feature.ranges:
                            yield (
                                feature_name,
                                sheet.title,
                                str(merged_range),
                                None,
                                None,
                            )
                    else:
                        yield feature_name, sheet.title, None, None, None
            except Exception as error:
                yield feature_name, sheet.title, None, None, str(error)


def _external_link_targets(workbook, path):
    targets = []
    for external_link in getattr(workbook, "_external_links", []):
        file_link = getattr(external_link, "file_link", None)
        target = getattr(file_link, "Target", None)
        if target is None:
            target = getattr(file_link, "target", None)
        if target and target not in targets:
            targets.append(target)
    try:
        with ZipFile(path) as archive:
            names = set(archive.namelist())
            for name in sorted(names):
                if not (
                    name.startswith("xl/externalLinks/externalLink")
                    and name.endswith(".xml")
                ):
                    continue
                relationship_name = name.replace(
                    "xl/externalLinks/",
                    "xl/externalLinks/_rels/",
                ).replace(".xml", ".xml.rels")
                if relationship_name not in names:
                    continue
                relationships = archive.read(relationship_name)
                for match in re.findall(
                    rb'Target="([^"]+)"',
                    relationships,
                ):
                    target = match.decode("utf-8")
                    if target not in targets:
                        targets.append(target)
    except (OSError, ValueError):
        pass
    return targets


def _package_feature_present(path, feature_name):
    try:
        with ZipFile(path) as archive:
            names = archive.namelist()
            if feature_name == "charts":
                return any(name.startswith("xl/charts/") for name in names)
            if feature_name == "pivot_tables":
                return any(
                    name.startswith("xl/pivotTables/") for name in names
                )
            if feature_name == "external_links":
                return any(
                    name.startswith("xl/externalLinks/") for name in names
                )
            if feature_name == "named_ranges":
                return b"<definedName " in archive.read("xl/workbook.xml")
            if feature_name == "merged_cells":
                return any(
                    b"<mergeCells" in archive.read(name)
                    for name in names
                    if name.startswith("xl/worksheets/")
                    and name.endswith(".xml")
                )
    except (KeyError, OSError, ValueError):
        return False
    return False


def _normalise_header(value):
    return str(value).strip().casefold()


def _header_values(sheet, header_row, table, resolution):
    max_column = None
    boundary = table.data_boundary
    if boundary is not None and boundary.mode == "excel_table":
        excel_table = sheet.tables.get(boundary.table_name)
        if excel_table is not None:
            _min_column, _min_row, max_column, _max_row = range_boundaries(
                excel_table.ref
            )
    if max_column is None:
        cells = resolution.cells(sheet)
        if cells is not None:
            max_column = max(
                (
                    cell.column
                    for cell in cells
                    if cell.row == header_row and cell.value is not None
                ),
                default=0,
            )
        else:
            max_column = sheet.max_column
    return [
        sheet.cell(header_row, column).value
        for column in range(1, max_column + 1)
    ]


def _is_blank_value(value, blank_policy, null_tokens):
    if blank_policy == "none_only":
        return value is None
    if blank_policy == "blank_strings":
        return value is None or value == ""
    if blank_policy == "blank_strings_and_nulls":
        return value is None or value == "" or value in null_tokens
    return value is None


@dataclass
class _WorkbookResolution:
    workbook: object
    _cell_cache: dict = field(default_factory=dict)
    _shape_cache: dict = field(default_factory=dict)
    _header_cache: dict = field(default_factory=dict)

    def cells(self, sheet):
        key = sheet.title
        if key not in self._cell_cache:
            stored = getattr(sheet, "_cells", None)
            self._cell_cache[key] = (
                tuple(stored.values())
                if stored is not None
                else tuple(_read_only_cells(sheet))
            )
        cached = self._cell_cache[key]
        return cached

    def shape(self, value_sheet, formula_sheet):
        key = (value_sheet.title, formula_sheet.title)
        if key not in self._shape_cache:
            self._shape_cache[key] = _content_shape(
                value_sheet,
                formula_sheet,
                self,
                self,
            )
        return self._shape_cache[key]

    def headers(self, sheet, header_row, table):
        boundary = table.data_boundary
        table_name = (
            boundary.table_name
            if boundary is not None and boundary.mode == "excel_table"
            else None
        )
        key = (sheet.title, header_row, table_name)
        if key not in self._header_cache:
            self._header_cache[key] = _header_values(
                sheet,
                header_row,
                table,
                self,
            )
        return self._header_cache[key]


def _cells_until_limit(workbook, limit, resolution=None):
    count = 0
    for sheet in workbook.worksheets:
        cells = (
            resolution.cells(sheet)
            if resolution is not None
            else _stored_cells(sheet)
        )
        for _cell in cells:
            count += 1
            if count > limit:
                return count
    return None


def _stored_cells(sheet):
    cells = getattr(sheet, "_cells", None)
    if cells is not None:
        return cells.values()
    return (
        cell
        for row in sheet.iter_rows()
        for cell in row
    )


def _read_only_cells(sheet):
    with sheet._get_source() as source:
        parser = WorkSheetParser(
            source,
            sheet._shared_strings,
            data_only=sheet.parent.data_only,
            epoch=sheet.parent.epoch,
            date_formats=sheet.parent._date_formats,
            timedelta_formats=sheet.parent._timedelta_formats,
        )
        for _row_number, row in parser.parse():
            for cell in sheet._get_row(row):
                if cell.value is not None:
                    yield cell


def _comparison_coordinates(
    candidate_sheet,
    reference_sheet,
    candidate_resolution=None,
    reference_resolution=None,
):
    if (
        hasattr(candidate_sheet, "_cells")
        and hasattr(reference_sheet, "_cells")
    ):
        coordinates = {
            (cell.row, cell.column)
            for cell in candidate_resolution.cells(candidate_sheet)
        }
        coordinates.update(
            (cell.row, cell.column)
            for cell in reference_resolution.cells(reference_sheet)
        )
        return coordinates
    max_row = max(candidate_sheet.max_row, reference_sheet.max_row)
    max_column = max(
        candidate_sheet.max_column,
        reference_sheet.max_column,
    )
    return (
        (row_number, column_number)
        for row_number in range(1, max_row + 1)
        for column_number in range(1, max_column + 1)
    )


def _table_data_end_row(
    table,
    header_row,
    value_sheet,
    formula_sheet,
    logical_columns,
    value_resolution=None,
    formula_resolution=None,
):
    max_row = max(value_sheet.max_row, formula_sheet.max_row)
    boundary = table.data_boundary
    if boundary is None:
        return max_row
    if boundary.mode == "fixed_end_row" and boundary.end_row is not None:
        return min(max_row, boundary.end_row)
    if boundary.mode == "excel_table" and boundary.table_name:
        excel_table = value_sheet.tables.get(boundary.table_name)
        if excel_table is None:
            return None
        _min_column, _min_row, _max_column, max_table_row = (
            range_boundaries(excel_table.ref)
        )
        return max_table_row
    if boundary.mode == "last_non_empty_row" and boundary.columns:
        column_numbers = [
            logical_columns[column]
            for column in boundary.columns
            if column in logical_columns
        ]
        if hasattr(value_sheet, "_cells") and hasattr(formula_sheet, "_cells"):
            rows = [
                cell.row
                for cell in (
                    value_resolution.cells(value_sheet)
                    if value_resolution is not None
                    else _stored_cells(value_sheet)
                )
                if cell.column in column_numbers
                and cell.row > header_row
                and cell.value is not None
            ]
            rows.extend(
                cell.row
                for cell in (
                    formula_resolution.cells(formula_sheet)
                    if formula_resolution is not None
                    else _stored_cells(formula_sheet)
                )
                if cell.column in column_numbers
                and cell.row > header_row
                and cell.value is not None
            )
            return max(rows, default=header_row)
        for row_number in range(max_row, header_row, -1):
            if any(
                value_sheet.cell(row_number, column).value is not None
                or formula_sheet.cell(row_number, column).value is not None
                for column in column_numbers
            ):
                return row_number
        return header_row
    return max_row


def _content_shape(
    value_sheet,
    formula_sheet,
    value_resolution=None,
    formula_resolution=None,
):
    max_row = 0
    max_column = 0
    if hasattr(value_sheet, "_cells") and hasattr(formula_sheet, "_cells"):
        coordinates = {
            cell.coordinate
            for cell in (
                value_resolution.cells(value_sheet)
                if value_resolution is not None
                else _stored_cells(value_sheet)
            )
        }
        coordinates.update(
            cell.coordinate
            for cell in (
                formula_resolution.cells(formula_sheet)
                if formula_resolution is not None
                else _stored_cells(formula_sheet)
            )
        )
        for coordinate in coordinates:
            value_cell = value_sheet[coordinate]
            formula_cell = formula_sheet[coordinate]
            if value_cell.value is None and formula_cell.value is None:
                continue
            max_row = max(max_row, value_cell.row, formula_cell.row)
            max_column = max(
                max_column,
                value_cell.column,
                formula_cell.column,
            )
        return {"max_row": max_row, "max_column": max_column}
    scan_max_row = max(value_sheet.max_row, formula_sheet.max_row)
    scan_max_column = max(value_sheet.max_column, formula_sheet.max_column)
    for row_number in range(1, scan_max_row + 1):
        for column_number in range(1, scan_max_column + 1):
            value_cell = value_sheet.cell(row_number, column_number)
            formula_cell = formula_sheet.cell(row_number, column_number)
            if value_cell.value is None and formula_cell.value is None:
                continue
            max_row = max(max_row, row_number)
            max_column = max(max_column, column_number)
    return {"max_row": max_row, "max_column": max_column}
